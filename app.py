"""
app.py  —  ProctorAI v6  (COCO YOLOv8 Final Edition)
======================================================
YOLOv8 real-time detection:
  • Class  0  → person      (multi-person alert)
  • Class 67  → cell phone  (COCO pretrained)
  • Class  1  → cell_phone  (fine-tuned yolo_proctor.pt)

5-Class CNN Detection:
  0 Normal | 1 Phone | 2 Multi-Person | 3 Voice | 4 Identity-Mismatch

Run:
    pip install flask werkzeug numpy opencv-python ultralytics openpyxl
    python app.py
    Student → http://127.0.0.1:5000/student/login
    Admin   → http://127.0.0.1:5000/admin/login   admin/admin123
"""

PORT       = 5000
SECRET_KEY = "cognizant_proctor_v6_2025"
MODEL_DIR  = "models"
SS_DIR     = "screenshots"
EXPORT_DIR = "exports"


EMAIL_CONFIG = {
    # provider: brevo | sendgrid | mailgun | gmail | outlook | yahoo | custom
    "provider":    "brevo",
    "api_key":     "",          # For Brevo / SendGrid / Mailgun — paste API key here
    "smtp_host":   "smtp-relay.brevo.com",
    "smtp_port":   587,
    "smtp_user":   "",          # For Brevo: your Brevo account email
    "smtp_pass":   "",          # For Brevo: SMTP Password from Brevo (not account password)
    "from_name":   "ProctorAI Exams",
    "from_email":  "",          # Sender email shown to students
    "use_ssl":     False,
    "enabled":     False,
}
EXAM_BASE_URL = "http://127.0.0.1:5000"

# Provider presets — host/port auto-filled when admin selects provider
PROVIDER_PRESETS = {
    "brevo":     {"smtp_host":"smtp-relay.brevo.com",  "smtp_port":587, "use_ssl":False,
                  "api_mode":True,
                  "help":"FREE 300 emails/day. Steps: 1) Create free account at brevo.com  2) Go to SMTP & API → SMTP  3) Copy the SMTP Login (email) and SMTP Password  4) Paste them above. No App Password needed!"},
    "sendgrid":  {"smtp_host":"smtp.sendgrid.net",     "smtp_port":587, "use_ssl":False,
                  "api_mode":True,
                  "help":"FREE 100 emails/day. Steps: 1) Create account at sendgrid.com  2) Go to Settings → API Keys → Create API Key (Full Access)  3) Use 'apikey' as username and paste the API key as password."},
    "mailgun":   {"smtp_host":"smtp.mailgun.org",      "smtp_port":587, "use_ssl":False,
                  "api_mode":True,
                  "help":"FREE 1000 emails/month. Steps: 1) Create account at mailgun.com  2) Go to Sending → Domains  3) Get SMTP credentials from your domain settings."},
    "gmail":     {"smtp_host":"smtp.gmail.com",        "smtp_port":587, "use_ssl":False,
                  "api_mode":False,
                  "help":"Requires Gmail App Password (NOT your login password). Enable 2FA → Google Account → Security → App Passwords → Generate."},
    "outlook":   {"smtp_host":"smtp.office365.com",    "smtp_port":587, "use_ssl":False,
                  "api_mode":False,
                  "help":"Use your Microsoft 365 / Outlook password or App Password."},
    "yahoo":     {"smtp_host":"smtp.mail.yahoo.com",   "smtp_port":587, "use_ssl":False,
                  "api_mode":False,
                  "help":"Create an App Password at Yahoo Account Security settings."},
    "custom":    {"smtp_host":"",                      "smtp_port":587, "use_ssl":False,
                  "api_mode":False,
                  "help":"Enter your company or college SMTP server details. Ask your IT team for: Host, Port, Username, Password."},
}

# ── Exam question bank config ───────────────────────────────────────────────
MCQ_PER_EXAM     = 20   # how many MCQ each student gets (from 480 bank)
SQL_PER_EXAM     = 2    # how many SQL questions each student gets
CODE_PER_EXAM    = 3    # how many coding questions each student gets (from 29 bank)
EXAM_LINK_EXPIRY = 24   # hours before link expires

# Supported compiler languages
SUPPORTED_LANGS = ["python", "java", "c", "cpp", "javascript"]

import os, sys, json, pickle, base64, time, uuid, csv, traceback
import subprocess, tempfile, sqlite3, smtplib, threading, secrets, hashlib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from functools import wraps

try:
    from flask import (Flask, render_template, request, jsonify, session,
                       redirect, url_for, send_file, send_from_directory)
    from werkzeug.security import generate_password_hash, check_password_hash
except ImportError:
    print("[ERROR] pip install flask werkzeug"); sys.exit(1)

import numpy as np

try:
    import tensorflow as tf
    HAS_TF = True
    print(f"[TF] TensorFlow {tf.__version__}")
except Exception:
    HAS_TF = False
    print("[TF] Not installed — CNN disabled. pip install tensorflow")

try:
    import cv2
    HAS_CV = True
    print(f"[CV2] OpenCV {cv2.__version__}")
except Exception:
    HAS_CV = False
    print("[CV2] Not installed — pip install opencv-python")

try:
    from ultralytics import YOLO as UltralyticsYOLO
    HAS_YOLO = True
    print("[YOLO] ultralytics ready")
except Exception:
    HAS_YOLO = False
    print("[YOLO] Not installed — pip install ultralytics")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    HAS_XL = True
except Exception:
    HAS_XL = False

for d in [MODEL_DIR, SS_DIR, EXPORT_DIR, "templates", "static"]:
    os.makedirs(d, exist_ok=True)

app = Flask(__name__, template_folder="templates")
app.secret_key = SECRET_KEY

ADMINS = {
    "admin":     generate_password_hash("admin123"),
    "cognizant": generate_password_hash("proctor2024"),
}

SESSIONS: dict = {}
_smooth_buf: dict = {}
_SMOOTH = {"phone": 2, "multi": 4, "voice": 1, "identity": 3}

# Termination limits
PHONE_TERMINATE         = 3   # 3 phone detections  → terminate
MULTI_TERMINATE         = 3   # 3 multi-person       → terminate
FACE_MISMATCH_TERMINATE = 3   # 3 face mismatches   → terminate
TAB_TERMINATE           = 2   # 2 tab switches       → terminate

NUM_CLASSES  = 5
NUM_FEATURES = 72
CLASS_NAMES  = ["Normal","Phone","Multi-Person","Voice","Identity-Mismatch"]
EXAM_DURATION = 1800   # seconds

# COCO class IDs
YOLO_PERSON_COCO = 0
YOLO_PHONE_COCO  = 67
# Fine-tuned class IDs (from train_yolo_coco.py)
YOLO_PERSON_FINE = 0
YOLO_PHONE_FINE  = 1

YOLO_PERSON_ID = 0
YOLO_PHONE_ID  = 67

# Model globals
_cnn_model    = None
_rf_model     = None
_scaler       = None
_yolo_model   = None
_model_type   = "heuristic"
_yolo_fine    = False

# ==============================================================================
# QUESTION BANK  (3 Parts: MCQ / SQL / Coding)
# Each student gets a RANDOM subset — different for every person
# ==============================================================================

# ── Part A: Multiple Choice Questions ─────────────────────────────────────────
MCQ_BANK = [
    {"id":"na_ns_1","question":"What is the LCM of 12 and 18?","options":["24", "36", "48", "60"],"correct":1,"marks":1,"topic":"Number System","section":"Numerical Ability"},
    {"id":"na_ns_2","question":"What is the HCF of 48 and 64?","options":["8", "12", "16", "24"],"correct":2,"marks":1,"topic":"Number System","section":"Numerical Ability"},
    {"id":"na_ns_3","question":"Which of the following is divisible by 11?","options":["121", "123", "125", "127"],"correct":0,"marks":1,"topic":"Divisibility","section":"Numerical Ability"},
    {"id":"na_ns_4","question":"What is the value of 0.75 as a fraction?","options":["3/5", "3/4", "2/3", "4/5"],"correct":1,"marks":1,"topic":"Decimal Fractions","section":"Numerical Ability"},
    {"id":"na_ns_5","question":"Find the LCM of 6, 8, and 12.","options":["24", "36", "48", "60"],"correct":0,"marks":1,"topic":"Number System","section":"Numerical Ability"},
    {"id":"na_ns_6","question":"A number is divisible by 9 if the sum of its digits is divisible by:","options":["3", "6", "9", "18"],"correct":2,"marks":1,"topic":"Divisibility","section":"Numerical Ability"},
    {"id":"na_ns_7","question":"What is 2.5 \u00d7 0.4?","options":["0.1", "1", "10", "0.01"],"correct":1,"marks":1,"topic":"Decimal Fractions","section":"Numerical Ability"},
    {"id":"na_ns_8","question":"The HCF of two numbers is 12 and their LCM is 144. If one number is 36, the other is:","options":["36", "48", "72", "24"],"correct":1,"marks":1,"topic":"Number System","section":"Numerical Ability"},
    {"id":"na_ns_9","question":"Which is a prime number?","options":["51", "57", "59", "63"],"correct":2,"marks":1,"topic":"Number System","section":"Numerical Ability"},
    {"id":"na_ns_10","question":"0.001 \u00d7 100 = ?","options":["0.1", "1", "10", "0.01"],"correct":0,"marks":1,"topic":"Decimal Fractions","section":"Numerical Ability"},
    {"id":"na_ns_11","question":"The LCM of two coprime numbers is always equal to:","options":["Their HCF", "Sum of numbers", "Their product", "Difference"],"correct":2,"marks":1,"topic":"Number System","section":"Numerical Ability"},
    {"id":"na_ns_12","question":"Which number is divisible by both 4 and 6?","options":["14", "18", "20", "24"],"correct":3,"marks":1,"topic":"Divisibility","section":"Numerical Ability"},
    {"id":"na_ns_13","question":"Express 7/8 as a decimal.","options":["0.785", "0.875", "0.758", "0.857"],"correct":1,"marks":1,"topic":"Decimal Fractions","section":"Numerical Ability"},
    {"id":"na_ns_14","question":"Find the HCF of 16, 24, and 40.","options":["4", "8", "6", "2"],"correct":1,"marks":1,"topic":"Number System","section":"Numerical Ability"},
    {"id":"na_ns_15","question":"A number when divided by 5 leaves remainder 3, when divided by 6 leaves remainder 4. The number is:","options":["28", "23", "33", "43"],"correct":1,"marks":1,"topic":"Number System","section":"Numerical Ability"},
    {"id":"na_ns_16","question":"What is 1.44 \u00d7 2.5?","options":["3.6", "3.2", "4.0", "2.8"],"correct":0,"marks":1,"topic":"Decimal Fractions","section":"Numerical Ability"},
    {"id":"na_ns_17","question":"How many prime numbers are there between 1 and 20?","options":["6", "7", "8", "9"],"correct":2,"marks":1,"topic":"Number System","section":"Numerical Ability"},
    {"id":"na_ns_18","question":"Find the LCM of 15 and 25.","options":["75", "50", "100", "125"],"correct":0,"marks":1,"topic":"Number System","section":"Numerical Ability"},
    {"id":"na_ns_19","question":"A number divisible by 8 must be divisible by:","options":["2 and 4", "3 and 4", "2 only", "4 only"],"correct":0,"marks":1,"topic":"Divisibility","section":"Numerical Ability"},
    {"id":"na_ns_20","question":"What is 3/7 \u00d7 14/9?","options":["2/3", "6/7", "1/3", "2/9"],"correct":0,"marks":1,"topic":"Number System","section":"Numerical Ability"},
    {"id":"na_men_1","question":"Area of a rectangle with length 8 and width 5 is:","options":["35", "40", "45", "50"],"correct":1,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_men_2","question":"Perimeter of a square with side 6 cm is:","options":["24 cm", "36 cm", "12 cm", "18 cm"],"correct":0,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_men_3","question":"Area of a circle with radius 7 (\u03c0=22/7) is:","options":["154", "144", "164", "134"],"correct":0,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_men_4","question":"Volume of a cube with side 4 cm is:","options":["16", "32", "64", "48"],"correct":2,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_men_5","question":"The diagonal of a square with side 5 cm is:","options":["5√2 cm", "10 cm", "5 cm", "25 cm"],"correct":0,"marks":1,"topic":"Geometry","section":"Numerical Ability"},
    {"id":"na_men_6","question":"Area of a triangle with base 10 and height 6 is:","options":["60", "30", "20", "40"],"correct":1,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_men_7","question":"Circumference of a circle with diameter 14 (\u03c0=22/7):","options":["44 cm", "22 cm", "88 cm", "33 cm"],"correct":0,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_men_8","question":"Surface area of a cube with side 3 cm:","options":["27 sq cm", "54 sq cm", "18 sq cm", "36 sq cm"],"correct":1,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_men_9","question":"In a right triangle, if legs are 3 and 4, hypotenuse is:","options":["5", "6", "7", "8"],"correct":0,"marks":1,"topic":"Geometry","section":"Numerical Ability"},
    {"id":"na_men_10","question":"Perimeter of a triangle with sides 5, 7, and 9:","options":["21", "20", "22", "19"],"correct":0,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_men_11","question":"Volume of a cylinder with r=7, h=10 (\u03c0=22/7):","options":["1540", "1440", "1640", "1240"],"correct":0,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_men_12","question":"Area of trapezium with parallel sides 6,10 and height 4:","options":["32", "64", "16", "48"],"correct":0,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_men_13","question":"Sum of angles in a triangle is:","options":["180°", "270°", "360°", "90°"],"correct":0,"marks":1,"topic":"Geometry","section":"Numerical Ability"},
    {"id":"na_men_14","question":"Radius of a circle with area 616 sq cm (\u03c0=22/7):","options":["7 cm", "14 cm", "21 cm", "28 cm"],"correct":1,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_men_15","question":"Area of a rhombus with diagonals 10 and 8:","options":["40", "80", "20", "60"],"correct":0,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_men_16","question":"The number of faces of a cuboid is:","options":["6", "8", "4", "12"],"correct":0,"marks":1,"topic":"Geometry","section":"Numerical Ability"},
    {"id":"na_men_17","question":"Perimeter of a semicircle with radius 7 (\u03c0=22/7):","options":["36 cm", "22 cm", "11 cm", "44 cm"],"correct":0,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_men_18","question":"Volume of a sphere with r=3 (\u03c0=22/7, approx):","options":["113.14", "108", "125", "100"],"correct":0,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_men_19","question":"An equilateral triangle has all sides equal and each angle equal to:","options":["60°", "90°", "45°", "120°"],"correct":0,"marks":1,"topic":"Geometry","section":"Numerical Ability"},
    {"id":"na_men_20","question":"Area of a parallelogram with base 12 and height 5:","options":["60", "120", "30", "55"],"correct":0,"marks":1,"topic":"Mensuration","section":"Numerical Ability"},
    {"id":"na_ar_1","question":"20% of 500 is:","options":["80", "100", "120", "60"],"correct":1,"marks":1,"topic":"Percentages","section":"Numerical Ability"},
    {"id":"na_ar_2","question":"A mixture has milk and water in ratio 3:2. In 25 litres, milk is:","options":["10 L", "15 L", "20 L", "12 L"],"correct":1,"marks":1,"topic":"Allegations and Mixtures","section":"Numerical Ability"},
    {"id":"na_ar_3","question":"If A:B = 2:3 and B:C = 4:5, then A:C =","options":["8:15", "2:5", "4:9", "6:10"],"correct":0,"marks":1,"topic":"Ratios","section":"Numerical Ability"},
    {"id":"na_ar_4","question":"A does a job in 10 days, B in 15 days. Together they finish in:","options":["6 days", "8 days", "5 days", "12 days"],"correct":0,"marks":1,"topic":"Work and Time","section":"Numerical Ability"},
    {"id":"na_ar_5","question":"Speed = Distance/Time. If distance is 120 km and time is 3 hours, speed is:","options":["40 km/h", "30 km/h", "50 km/h", "60 km/h"],"correct":0,"marks":1,"topic":"Speed Time Distance","section":"Numerical Ability"},
    {"id":"na_ar_6","question":"CP=200, SP=250. Profit% is:","options":["20%", "25%", "30%", "15%"],"correct":1,"marks":1,"topic":"Profit and Loss","section":"Numerical Ability"},
    {"id":"na_ar_7","question":"Average of 5, 10, 15, 20, 25 is:","options":["15", "16", "14", "17"],"correct":0,"marks":1,"topic":"Averages","section":"Numerical Ability"},
    {"id":"na_ar_8","question":"Solve: 2x + 5 = 15","options":["x=4", "x=5", "x=6", "x=3"],"correct":1,"marks":1,"topic":"Equations","section":"Numerical Ability"},
    {"id":"na_ar_9","question":"Father's age is 4\u00d7 son's age. In 5 years, father will be 3\u00d7 son's age. Son's age now?","options":["5", "10", "15", "20"],"correct":1,"marks":1,"topic":"Ages","section":"Numerical Ability"},
    {"id":"na_ar_10","question":"Next term in series: 2, 6, 12, 20, 30, ?","options":["42", "40", "44", "36"],"correct":0,"marks":1,"topic":"Series","section":"Numerical Ability"},
    {"id":"na_ar_11","question":"A shopkeeper gives 10% discount on marked price \u20b9500. SP is:","options":["₹450", "₹400", "₹480", "₹425"],"correct":0,"marks":1,"topic":"Profit and Loss","section":"Numerical Ability"},
    {"id":"na_ar_12","question":"In what ratio should water be added to spirit worth \u20b960/L to get a mixture worth \u20b950/L?","options":["1:5", "1:4", "1:3", "2:5"],"correct":0,"marks":1,"topic":"Allegations and Mixtures","section":"Numerical Ability"},
    {"id":"na_ar_13","question":"A train 200m long passes a pole in 10 seconds. Speed of train is:","options":["20 m/s", "18 m/s", "25 m/s", "22 m/s"],"correct":0,"marks":1,"topic":"Speed Time Distance","section":"Numerical Ability"},
    {"id":"na_ar_14","question":"What percent is 75 of 300?","options":["20%", "25%", "30%", "15%"],"correct":1,"marks":1,"topic":"Percentages","section":"Numerical Ability"},
    {"id":"na_ar_15","question":"15 workers finish work in 6 days. How many to finish in 9 days?","options":["10", "8", "12", "9"],"correct":0,"marks":1,"topic":"Work and Time","section":"Numerical Ability"},
    {"id":"na_ar_16","question":"Average of first 10 natural numbers is:","options":["5", "5.5", "6", "4.5"],"correct":1,"marks":1,"topic":"Averages","section":"Numerical Ability"},
    {"id":"na_ar_17","question":"The sum of three consecutive even numbers is 72. Largest is:","options":["26", "28", "24", "30"],"correct":0,"marks":1,"topic":"Equations","section":"Numerical Ability"},
    {"id":"na_ar_18","question":"Ratio of boys to girls is 3:5. If 40 girls, number of boys:","options":["24", "20", "30", "15"],"correct":0,"marks":1,"topic":"Ratios","section":"Numerical Ability"},
    {"id":"na_ar_19","question":"Sum of series 1+2+3+...+20 is:","options":["210", "200", "190", "220"],"correct":0,"marks":1,"topic":"Series","section":"Numerical Ability"},
    {"id":"na_ar_20","question":"A person's salary increased by 20% then decreased by 20%. Net change?","options":["0%", "4% decrease", "4% increase", "2% decrease"],"correct":1,"marks":1,"topic":"Percentages","section":"Numerical Ability"},
    {"id":"na_st_1","question":"Mean of 4, 8, 6, 10, 2 is:","options":["5", "6", "7", "8"],"correct":1,"marks":1,"topic":"Mean","section":"Numerical Ability"},
    {"id":"na_st_2","question":"Median of 3, 7, 1, 9, 5 is:","options":["5", "7", "3", "6"],"correct":0,"marks":1,"topic":"Median","section":"Numerical Ability"},
    {"id":"na_st_3","question":"Mode of 2, 3, 3, 5, 7, 3 is:","options":["2", "3", "5", "7"],"correct":1,"marks":1,"topic":"Mode","section":"Numerical Ability"},
    {"id":"na_st_4","question":"Variance is the square of:","options":["Mean", "Median", "Standard Deviation", "Mode"],"correct":2,"marks":1,"topic":"Variance","section":"Numerical Ability"},
    {"id":"na_st_5","question":"Standard Deviation of 2, 4, 4, 4, 5, 5, 7, 9 is:","options":["1", "2", "3", "4"],"correct":1,"marks":1,"topic":"Standard Deviation","section":"Numerical Ability"},
    {"id":"na_st_6","question":"If mean is 10 and SD is 2, coefficient of variation is:","options":["20%", "15%", "25%", "10%"],"correct":0,"marks":1,"topic":"Standard Deviation","section":"Numerical Ability"},
    {"id":"na_st_7","question":"Median of even number of values is:","options":["Middle value", "Average of two middle values", "Largest value", "Sum/Count"],"correct":1,"marks":1,"topic":"Median","section":"Numerical Ability"},
    {"id":"na_st_8","question":"Mean of first 5 odd numbers is:","options":["5", "4", "6", "3"],"correct":0,"marks":1,"topic":"Mean","section":"Numerical Ability"},
    {"id":"na_st_9","question":"If all values are equal, standard deviation is:","options":["0", "1", "Mean value", "Undefined"],"correct":0,"marks":1,"topic":"Standard Deviation","section":"Numerical Ability"},
    {"id":"na_st_10","question":"Range = Maximum value \u2212 ?","options":["Mean", "Median", "Minimum value", "Mode"],"correct":2,"marks":1,"topic":"Variance","section":"Numerical Ability"},
    {"id":"na_st_11","question":"Mode of 1, 2, 2, 3, 4, 4, 4, 5 is:","options":["2", "4", "3", "1"],"correct":1,"marks":1,"topic":"Mode","section":"Numerical Ability"},
    {"id":"na_st_12","question":"The sum of deviations from mean is always:","options":["0", "1", "Positive", "Negative"],"correct":0,"marks":1,"topic":"Mean","section":"Numerical Ability"},
    {"id":"na_st_13","question":"Mean of 10 numbers is 15. Sum of all 10 numbers is:","options":["150", "50", "115", "25"],"correct":0,"marks":1,"topic":"Mean","section":"Numerical Ability"},
    {"id":"na_st_14","question":"Arrange in order for median of 7 numbers: take the __ value:","options":["3rd", "4th", "5th", "6th"],"correct":1,"marks":1,"topic":"Median","section":"Numerical Ability"},
    {"id":"na_st_15","question":"Variance of data: 2, 2, 2, 2, 2 is:","options":["0", "2", "4", "1"],"correct":0,"marks":1,"topic":"Variance","section":"Numerical Ability"},
    {"id":"na_st_16","question":"In positively skewed distribution: Mean __ Mode","options":["< ", "= ", "> ", "≤"],"correct":2,"marks":1,"topic":"Mean","section":"Numerical Ability"},
    {"id":"na_st_17","question":"Standard deviation is always:","options":["Negative", "Zero", "Non-negative", "Greater than mean"],"correct":2,"marks":1,"topic":"Standard Deviation","section":"Numerical Ability"},
    {"id":"na_st_18","question":"Which measure of central tendency is most affected by extreme values?","options":["Median", "Mode", "Mean", "All equally"],"correct":2,"marks":1,"topic":"Mean","section":"Numerical Ability"},
    {"id":"na_st_19","question":"If each value increased by 5, mean will:","options":["Increase by 5", "Stay same", "Decrease by 5", "Double"],"correct":0,"marks":1,"topic":"Mean","section":"Numerical Ability"},
    {"id":"na_st_20","question":"Interquartile range uses which percentiles?","options":["25th and 75th", "10th and 90th", "0 and 100", "50th only"],"correct":0,"marks":1,"topic":"Variance","section":"Numerical Ability"},
    {"id":"na_di_1","question":"A pie chart has a sector of 90\u00b0. What % does it represent?","options":["25%", "30%", "20%", "15%"],"correct":0,"marks":1,"topic":"Pie Charts","section":"Numerical Ability"},
    {"id":"na_di_2","question":"In a table, if sales in Jan=100, Feb=120, growth % in Feb is:","options":["10%", "20%", "15%", "25%"],"correct":1,"marks":1,"topic":"Tabular DI","section":"Numerical Ability"},
    {"id":"na_di_3","question":"A bar graph shows production: 2020=500, 2021=600. % increase?","options":["15%", "20%", "25%", "10%"],"correct":1,"marks":1,"topic":"Graphical DI","section":"Numerical Ability"},
    {"id":"na_di_4","question":"Pie chart sector = 72\u00b0. It represents what fraction?","options":["1/5", "1/4", "1/6", "1/3"],"correct":0,"marks":1,"topic":"Pie Charts","section":"Numerical Ability"},
    {"id":"na_di_5","question":"From a table: A scored 80, B scored 60. A scored what % more than B?","options":["25%", "33.3%", "20%", "40%"],"correct":1,"marks":1,"topic":"Tabular DI","section":"Numerical Ability"},
    {"id":"na_di_6","question":"A line graph shows temp Mon=20\u00b0C, Fri=25\u00b0C. Rise over 4 days?","options":["1.25°/day", "5°/day", "0.5°/day", "2°/day"],"correct":0,"marks":1,"topic":"Graphical DI","section":"Numerical Ability"},
    {"id":"na_di_7","question":"In a pie chart of 360\u00b0, if one category has 120\u00b0, its share is:","options":["1/3", "1/4", "1/2", "2/5"],"correct":0,"marks":1,"topic":"Pie Charts","section":"Numerical Ability"},
    {"id":"na_di_8","question":"Table shows expenses: Food=\u20b9500, Rent=\u20b91000, Others=\u20b9500. % on rent?","options":["50%", "40%", "25%", "33%"],"correct":0,"marks":1,"topic":"Tabular DI","section":"Numerical Ability"},
    {"id":"na_di_9","question":"Bar chart: 2018=200 units, 2022=400 units. What is CAGR (approx)?","options":["~19%", "~10%", "~15%", "~25%"],"correct":0,"marks":1,"topic":"Graphical DI","section":"Numerical Ability"},
    {"id":"na_di_10","question":"A pie chart shows 5 equal sectors. Each sector's degree?","options":["72°", "60°", "90°", "45°"],"correct":0,"marks":1,"topic":"Pie Charts","section":"Numerical Ability"},
    {"id":"na_di_11","question":"If total sales = 5000 and product A has 30%, A's sales =","options":["1500", "1000", "2000", "1200"],"correct":0,"marks":1,"topic":"Tabular DI","section":"Numerical Ability"},
    {"id":"na_di_12","question":"A histogram bar has width 5 and height 10. Area of bar =","options":["50", "15", "2", "55"],"correct":0,"marks":1,"topic":"Graphical DI","section":"Numerical Ability"},
    {"id":"na_di_13","question":"From table: Jan profit=200, Feb profit=150. Feb decline %?","options":["25%", "20%", "33%", "15%"],"correct":0,"marks":1,"topic":"Tabular DI","section":"Numerical Ability"},
    {"id":"na_di_14","question":"Pie chart: total = \u20b912000. Sector = 60\u00b0. Value of sector?","options":["₹2000", "₹3000", "₹1500", "₹4000"],"correct":0,"marks":1,"topic":"Pie Charts","section":"Numerical Ability"},
    {"id":"na_di_15","question":"Stacked bar: Category A=40%, B=35%, C=25%. Which is smallest?","options":["C", "A", "B", "All equal"],"correct":0,"marks":1,"topic":"Graphical DI","section":"Numerical Ability"},
    {"id":"na_di_16","question":"A table shows 4 students' marks. Average of 90,80,70,60 is:","options":["75", "80", "70", "85"],"correct":0,"marks":1,"topic":"Tabular DI","section":"Numerical Ability"},
    {"id":"na_di_17","question":"Double bar chart compares data of how many groups?","options":["2", "3", "4", "1"],"correct":0,"marks":1,"topic":"Graphical DI","section":"Numerical Ability"},
    {"id":"na_di_18","question":"In a pie chart, if sector A=144\u00b0 and B=72\u00b0, ratio A:B =","options":["2:1", "1:2", "3:1", "1:3"],"correct":0,"marks":1,"topic":"Pie Charts","section":"Numerical Ability"},
    {"id":"na_di_19","question":"From a line graph, if value decreases from 100 to 80, % decrease?","options":["20%", "25%", "15%", "30%"],"correct":0,"marks":1,"topic":"Graphical DI","section":"Numerical Ability"},
    {"id":"na_di_20","question":"Table: 5 items cost \u20b910,\u20b920,\u20b930,\u20b940,\u20b950. Total =","options":["₹150", "₹100", "₹200", "₹125"],"correct":0,"marks":1,"topic":"Tabular DI","section":"Numerical Ability"},
    {"id":"na_sim_1","question":"Simplify: 25 \u00d7 4 \u00f7 5 + 10","options":["30", "20", "40", "50"],"correct":2,"marks":1,"topic":"Simplification","section":"Numerical Ability"},
    {"id":"na_sim_2","question":"\u221a144 + \u221a81 =","options":["21", "20", "18", "15"],"correct":0,"marks":1,"topic":"Simplification","section":"Numerical Ability"},
    {"id":"na_sim_3","question":"Approximate: 19.98 \u00d7 5.01 \u2248","options":["100", "200", "150", "50"],"correct":0,"marks":1,"topic":"Approximation","section":"Numerical Ability"},
    {"id":"na_sim_4","question":"3\u00b3 + 4\u00b2 \u2212 2\u00b9 =","options":["39", "41", "43", "45"],"correct":1,"marks":1,"topic":"Simplification","section":"Numerical Ability"},
    {"id":"na_sim_5","question":"(100 \u2212 30%) of 200 =","options":["140", "160", "120", "180"],"correct":0,"marks":1,"topic":"Approximation","section":"Numerical Ability"},
    {"id":"na_sim_6","question":"Simplify: 1/2 + 1/3 + 1/6","options":["1", "2/3", "5/6", "7/6"],"correct":0,"marks":1,"topic":"Simplification","section":"Numerical Ability"},
    {"id":"na_sim_7","question":"Approximate: 7.99\u00b2 \u2248","options":["64", "63", "65", "70"],"correct":0,"marks":1,"topic":"Approximation","section":"Numerical Ability"},
    {"id":"na_sim_8","question":"(15% of 400) + (25% of 200) =","options":["110", "100", "120", "90"],"correct":0,"marks":1,"topic":"Simplification","section":"Numerical Ability"},
    {"id":"na_sim_9","question":"Simplify: (16 \u00d7 3) \u00f7 (4 \u00d7 2)","options":["6", "3", "12", "8"],"correct":0,"marks":1,"topic":"Simplification","section":"Numerical Ability"},
    {"id":"na_sim_10","question":"Approximate: 998 \u00d7 1002 \u2248","options":["1000000", "100000", "99996", "998000"],"correct":0,"marks":1,"topic":"Approximation","section":"Numerical Ability"},
    {"id":"na_sim_11","question":"2\u2075 \u00d7 2\u00b3 = ?","options":["2⁸", "2¹⁵", "4⁸", "2⁷"],"correct":0,"marks":1,"topic":"Simplification","section":"Numerical Ability"},
    {"id":"na_sim_12","question":"\u221a(169) \u00d7 \u221a(25) =","options":["65", "45", "55", "75"],"correct":0,"marks":1,"topic":"Simplification","section":"Numerical Ability"},
    {"id":"na_sim_13","question":"Approx: 49.1% of 800 \u2248","options":["400", "350", "450", "300"],"correct":0,"marks":1,"topic":"Approximation","section":"Numerical Ability"},
    {"id":"na_sim_14","question":"(48 \u00f7 6) \u00d7 (5 + 3) \u2212 10 =","options":["54", "64", "44", "34"],"correct":0,"marks":1,"topic":"Simplification","section":"Numerical Ability"},
    {"id":"na_sim_15","question":"Simplify: 3/4 of 120 + 25%of 80","options":["110", "100", "120", "90"],"correct":0,"marks":1,"topic":"Simplification","section":"Numerical Ability"},
    {"id":"na_sim_16","question":"Approximate: 4.02 \u00d7 9.98 \u2248","options":["40", "36", "44", "50"],"correct":0,"marks":1,"topic":"Approximation","section":"Numerical Ability"},
    {"id":"na_sim_17","question":"(100)\u00b2 \u2212 (99)\u00b2 = ?","options":["199", "201", "99", "101"],"correct":0,"marks":1,"topic":"Simplification","section":"Numerical Ability"},
    {"id":"na_sim_18","question":"Simplify: 5! / 3!","options":["20", "30", "60", "10"],"correct":0,"marks":1,"topic":"Simplification","section":"Numerical Ability"},
    {"id":"na_sim_19","question":"Approximate: \u221a(2025) \u2248","options":["45", "40", "50", "55"],"correct":0,"marks":1,"topic":"Approximation","section":"Numerical Ability"},
    {"id":"na_sim_20","question":"(3 + 4)\u00b2 \u2212 (3\u00b2 + 4\u00b2) =","options":["24", "12", "10", "25"],"correct":0,"marks":1,"topic":"Simplification","section":"Numerical Ability"},
    {"id":"ra_wp_1","question":"CLOCK : TIME :: THERMOMETER : ?","options":["Heat", "Temperature", "Weather", "Pressure"],"correct":1,"marks":1,"topic":"Word Pattern","section":"Reasoning Ability"},
    {"id":"ra_wp_2","question":"Find the odd one out: Apple, Mango, Carrot, Banana","options":["Apple", "Mango", "Carrot", "Banana"],"correct":2,"marks":1,"topic":"Word Pattern","section":"Reasoning Ability"},
    {"id":"ra_wp_3","question":"Next in letter series: A, C, E, G, ?","options":["H", "I", "J", "K"],"correct":1,"marks":1,"topic":"Letter Series","section":"Reasoning Ability"},
    {"id":"ra_wp_4","question":"BDFH : ACEG :: JLNP : ?","options":["IKMO", "IMKO", "IOKM", "KIOM"],"correct":0,"marks":1,"topic":"Letter Series","section":"Reasoning Ability"},
    {"id":"ra_wp_5","question":"Odd one out: Pen, Pencil, Eraser, Book, Ink","options":["Pen", "Pencil", "Book", "Eraser"],"correct":2,"marks":1,"topic":"Word Pattern","section":"Reasoning Ability"},
    {"id":"ra_wp_6","question":"ACE : BDF :: GIK : ?","options":["HJL", "HIJ", "JKL", "IKM"],"correct":0,"marks":1,"topic":"Letter Series","section":"Reasoning Ability"},
    {"id":"ra_wp_7","question":"ROSE : FLOWER :: COBRA : ?","options":["Animal", "Reptile", "Snake", "Venom"],"correct":2,"marks":1,"topic":"Word Pattern","section":"Reasoning Ability"},
    {"id":"ra_wp_8","question":"Series: Z X V T R P ?","options":["O", "N", "M", "L"],"correct":1,"marks":1,"topic":"Letter Series","section":"Reasoning Ability"},
    {"id":"ra_wp_9","question":"Doctor : Hospital :: Teacher : ?","options":["Book", "School", "Student", "Lesson"],"correct":1,"marks":1,"topic":"Word Pattern","section":"Reasoning Ability"},
    {"id":"ra_wp_10","question":"Series: AZ, BY, CX, DW, ?","options":["EV", "FU", "EW", "EU"],"correct":0,"marks":1,"topic":"Letter Series","section":"Reasoning Ability"},
    {"id":"ra_wp_11","question":"ABCD : DCBA :: EFGH : ?","options":["HGFE", "FGHE", "HFGE", "GHEF"],"correct":0,"marks":1,"topic":"Letter Series","section":"Reasoning Ability"},
    {"id":"ra_wp_12","question":"Odd out: Jupiter, Mars, Sun, Saturn, Venus","options":["Jupiter", "Mars", "Sun", "Saturn"],"correct":2,"marks":1,"topic":"Word Pattern","section":"Reasoning Ability"},
    {"id":"ra_wp_13","question":"Series: AEI, BFJ, CGK, ?","options":["DHL", "DHM", "DIL", "EHL"],"correct":0,"marks":1,"topic":"Letter Series","section":"Reasoning Ability"},
    {"id":"ra_wp_14","question":"PAINT : ARTIST :: STETHOSCOPE : ?","options":["Patient", "Nurse", "Doctor", "Hospital"],"correct":2,"marks":1,"topic":"Word Pattern","section":"Reasoning Ability"},
    {"id":"ra_wp_15","question":"Series: MNO, PQR, STU, ?","options":["VWX", "WXY", "UVW", "XYZ"],"correct":0,"marks":1,"topic":"Letter Series","section":"Reasoning Ability"},
    {"id":"ra_wp_16","question":"Odd out: Wheat, Rice, Cotton, Maize, Millet","options":["Wheat", "Rice", "Cotton", "Maize"],"correct":2,"marks":1,"topic":"Word Pattern","section":"Reasoning Ability"},
    {"id":"ra_wp_17","question":"ABD : CFI :: EGJ : ?","options":["HKO", "ILO", "HLO", "IKO"],"correct":0,"marks":1,"topic":"Letter Series","section":"Reasoning Ability"},
    {"id":"ra_wp_18","question":"Chair : Furniture :: Poem : ?","options":["Song", "Literature", "Novel", "Story"],"correct":1,"marks":1,"topic":"Word Pattern","section":"Reasoning Ability"},
    {"id":"ra_wp_19","question":"ZYX, WVU, TSR, ?","options":["QPO", "OPQ", "PQR", "NOP"],"correct":0,"marks":1,"topic":"Letter Series","section":"Reasoning Ability"},
    {"id":"ra_wp_20","question":"BIRD : AVIARY :: LION : ?","options":["Jungle", "Den", "Forest", "Cage"],"correct":1,"marks":1,"topic":"Word Pattern","section":"Reasoning Ability"},
    {"id":"ra_np_1","question":"Number series: 2, 4, 8, 16, ?","options":["24", "32", "30", "28"],"correct":1,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_2","question":"Series: 1, 4, 9, 16, 25, ?","options":["30", "35", "36", "49"],"correct":2,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_3","question":"Series: 3, 6, 11, 18, 27, ?","options":["38", "37", "39", "40"],"correct":0,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_4","question":"Odd out: 8, 27, 64, 100, 125","options":["8", "27", "100", "125"],"correct":2,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_5","question":"Series: 1, 1, 2, 3, 5, 8, ?","options":["11", "12", "13", "14"],"correct":2,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_6","question":"Series: 100, 81, 64, 49, ?","options":["36", "25", "16", "9"],"correct":0,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_7","question":"Find next: 5, 10, 20, 40, ?","options":["60", "70", "80", "90"],"correct":2,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_8","question":"Series: 2, 5, 10, 17, 26, ?","options":["35", "37", "36", "38"],"correct":1,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_9","question":"Odd one out: 3, 7, 11, 14, 19","options":["3", "11", "14", "19"],"correct":2,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_10","question":"Series: 1000, 100, 10, ?","options":["1", "0.1", "5", "2"],"correct":0,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_11","question":"Series: 7, 14, 28, 56, ?","options":["100", "112", "108", "120"],"correct":1,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_12","question":"Series: 4, 7, 12, 19, 28, ?","options":["37", "39", "38", "40"],"correct":1,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_13","question":"Odd out: 2, 3, 5, 7, 9, 11","options":["2", "5", "9", "11"],"correct":2,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_14","question":"Series: 0, 2, 6, 12, 20, ?","options":["28", "30", "32", "34"],"correct":1,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_15","question":"Find missing: 8, ?, 24, 48, 96","options":["12", "16", "18", "14"],"correct":1,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_16","question":"Series: 1, 8, 27, 64, ?","options":["100", "125", "81", "216"],"correct":1,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_17","question":"Odd out: 144, 169, 196, 200, 225","options":["144", "169", "200", "225"],"correct":2,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_18","question":"Series: 3, 9, 27, 81, ?","options":["162", "243", "189", "210"],"correct":1,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_19","question":"Series: 11, 13, 17, 19, 23, 29, ?","options":["31", "33", "37", "39"],"correct":0,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_np_20","question":"Find: 1+2+3+...+10 = ? Pattern","options":["50", "55", "45", "60"],"correct":1,"marks":1,"topic":"Number Series","section":"Reasoning Ability"},
    {"id":"ra_ps_1","question":"In a row of 10 people, A is 3rd from left. Position from right?","options":["7th", "8th", "6th", "9th"],"correct":1,"marks":1,"topic":"Seating Arrangements","section":"Reasoning Ability"},
    {"id":"ra_ps_2","question":"If MANGO is coded as NBNHP, how is APPLE coded?","options":["BQQMF", "BQPMF", "BQQNF", "BQQME"],"correct":0,"marks":1,"topic":"Coding-Decoding","section":"Reasoning Ability"},
    {"id":"ra_ps_3","question":"A is B's father. B is C's sister. D is C's mother. How is A related to D?","options":["Son", "Brother", "Husband", "Father"],"correct":2,"marks":1,"topic":"Blood Relations","section":"Reasoning Ability"},
    {"id":"ra_ps_4","question":"I walk 5 km North, turn right walk 3 km, turn right walk 5 km. Where am I from start?","options":["3 km West", "3 km East", "5 km South", "3 km North"],"correct":1,"marks":1,"topic":"Directional Sense","section":"Reasoning Ability"},
    {"id":"ra_ps_5","question":"5 persons sit in a row. A is at one end, E is next to A. B is in the middle. C is between D and E. Who is at other end?","options":["D", "C", "B", "E"],"correct":0,"marks":1,"topic":"Seating Arrangements","section":"Reasoning Ability"},
    {"id":"ra_ps_6","question":"If CAT = 3+1+20=24, DOG = ?","options":["26", "27", "28", "25"],"correct":1,"marks":1,"topic":"Coding-Decoding","section":"Reasoning Ability"},
    {"id":"ra_ps_7","question":"P is Q's brother. R is Q's mother. S is R's father. T is S's wife. How is T related to P?","options":["Mother", "Grandmother", "Aunt", "Sister"],"correct":1,"marks":1,"topic":"Blood Relations","section":"Reasoning Ability"},
    {"id":"ra_ps_8","question":"Facing North, turn 90\u00b0 clockwise, then 180\u00b0. Now facing?","options":["South", "East", "West", "North"],"correct":2,"marks":1,"topic":"Directional Sense","section":"Reasoning Ability"},
    {"id":"ra_ps_9","question":"6 people sit around a circular table. A is opposite B. How many people between A and B?","options":["2", "3", "1", "0"],"correct":0,"marks":1,"topic":"Seating Arrangements","section":"Reasoning Ability"},
    {"id":"ra_ps_10","question":"If BOOK is coded as CPPL, CODE is?","options":["DPEF", "DPCF", "EPEF", "DPFF"],"correct":0,"marks":1,"topic":"Coding-Decoding","section":"Reasoning Ability"},
    {"id":"ra_ps_11","question":"X is Y's husband. Y is Z's mother. Z is W's sister. How is X related to W?","options":["Uncle", "Father", "Brother", "Grandfather"],"correct":1,"marks":1,"topic":"Blood Relations","section":"Reasoning Ability"},
    {"id":"ra_ps_12","question":"Walk 10m East, 6m North, 10m West. How far from start?","options":["6m", "10m", "16m", "4m"],"correct":0,"marks":1,"topic":"Directional Sense","section":"Reasoning Ability"},
    {"id":"ra_ps_13","question":"A>B>C>D>E. Who is second smallest?","options":["D", "C", "B", "E"],"correct":0,"marks":1,"topic":"Seating Arrangements","section":"Reasoning Ability"},
    {"id":"ra_ps_14","question":"If in a code, RAIN = 4152, WIND = ?","options":["9254", "5329", "5294", "2945"],"correct":2,"marks":1,"topic":"Coding-Decoding","section":"Reasoning Ability"},
    {"id":"ra_ps_15","question":"A man pointing to a photo says 'she is the daughter of my grandfather's only son'. Relation?","options":["Sister", "Daughter", "Wife", "Niece"],"correct":0,"marks":1,"topic":"Blood Relations","section":"Reasoning Ability"},
    {"id":"ra_ps_16","question":"Sun sets in the West. It is evening and shadow falls East. I face my shadow. I face:","options":["West", "East", "North", "South"],"correct":0,"marks":1,"topic":"Directional Sense","section":"Reasoning Ability"},
    {"id":"ra_ps_17","question":"8 students sit in 2 rows of 4. How many are not at corners?","options":["4", "2", "6", "0"],"correct":0,"marks":1,"topic":"Seating Arrangements","section":"Reasoning Ability"},
    {"id":"ra_ps_18","question":"If 1=5, 2=25, 3=125, 4=625, then 5=?","options":["1", "3125", "5", "3025"],"correct":0,"marks":1,"topic":"Coding-Decoding","section":"Reasoning Ability"},
    {"id":"ra_ps_19","question":"A is B's mother's mother's husband. How is A related to B?","options":["Father", "Grandfather", "Uncle", "Paternal uncle"],"correct":1,"marks":1,"topic":"Blood Relations","section":"Reasoning Ability"},
    {"id":"ra_ps_20","question":"A walks 3 km South, 4 km East. Distance from starting point?","options":["5 km", "7 km", "6 km", "4 km"],"correct":0,"marks":1,"topic":"Directional Sense","section":"Reasoning Ability"},
    {"id":"ra_fig_1","question":"A cube has 6 faces. When unfolded, it forms a cross pattern with how many squares?","options":["6", "4", "8", "5"],"correct":0,"marks":1,"topic":"Cube Folding","section":"Reasoning Ability"},
    {"id":"ra_fig_2","question":"A square paper is folded diagonally once and cut at corner. When unfolded, holes are:","options":["1", "2", "4", "3"],"correct":2,"marks":1,"topic":"Paper Cuts","section":"Reasoning Ability"},
    {"id":"ra_fig_3","question":"Opposite faces of a cube sum to 7. Face with 3 is opposite to:","options":["4", "5", "6", "2"],"correct":0,"marks":1,"topic":"Cube Folding","section":"Reasoning Ability"},
    {"id":"ra_fig_4","question":"Paper folded in half and a hole punched. When unfolded, holes =","options":["1", "2", "3", "4"],"correct":1,"marks":1,"topic":"Paper Cuts","section":"Reasoning Ability"},
    {"id":"ra_fig_5","question":"A cube painted red on all faces is cut into 27 smaller cubes. Cubes with no paint?","options":["1", "8", "6", "0"],"correct":0,"marks":1,"topic":"Cube Folding","section":"Reasoning Ability"},
    {"id":"ra_fig_6","question":"Rectangle folded vertically, then horizontally, then cut diagonally. Pieces when unfolded?","options":["4", "8", "2", "16"],"correct":1,"marks":1,"topic":"Paper Cuts","section":"Reasoning Ability"},
    {"id":"ra_fig_7","question":"Which face of a cube is opposite to the top?","options":["Bottom", "Front", "Back", "Left"],"correct":0,"marks":1,"topic":"Cube Folding","section":"Reasoning Ability"},
    {"id":"ra_fig_8","question":"A square paper folded twice and punched once. Holes = ?","options":["4", "2", "1", "8"],"correct":0,"marks":1,"topic":"Paper Cuts","section":"Reasoning Ability"},
    {"id":"ra_fig_9","question":"In dice, opposite faces add to 7. If 1 is opposite 6, 2 is opposite?","options":["5", "4", "3", "6"],"correct":0,"marks":1,"topic":"Cube Folding","section":"Reasoning Ability"},
    {"id":"ra_fig_10","question":"Paper is folded to form a triangle from rectangle. Number of layers at fold?","options":["2", "3", "4", "1"],"correct":0,"marks":1,"topic":"Paper Cuts","section":"Reasoning Ability"},
    {"id":"ra_fig_11","question":"A cube cut into 64 equal pieces. Cuts needed on each axis:","options":["3", "4", "2", "5"],"correct":1,"marks":1,"topic":"Cube Folding","section":"Reasoning Ability"},
    {"id":"ra_fig_12","question":"Paper folded 3 times. Holes punched = 1. Unfolded holes = ?","options":["8", "4", "6", "2"],"correct":0,"marks":1,"topic":"Paper Cuts","section":"Reasoning Ability"},
    {"id":"ra_fig_13","question":"Cube painted 2 adjacent faces. Cut into 8 cubes. Cubes with 2 painted faces:","options":["1", "2", "4", "0"],"correct":0,"marks":1,"topic":"Cube Folding","section":"Reasoning Ability"},
    {"id":"ra_fig_14","question":"A+shaped cut on folded paper creates how many triangles when unfolded?","options":["4", "8", "2", "6"],"correct":0,"marks":1,"topic":"Paper Cuts","section":"Reasoning Ability"},
    {"id":"ra_fig_15","question":"On a standard dice, sum of all faces is:","options":["21", "18", "24", "15"],"correct":0,"marks":1,"topic":"Cube Folding","section":"Reasoning Ability"},
    {"id":"ra_fig_16","question":"A square paper folded on diagonal \u2014 shape formed?","options":["Triangle", "Rectangle", "Pentagon", "Hexagon"],"correct":0,"marks":1,"topic":"Paper Cuts","section":"Reasoning Ability"},
    {"id":"ra_fig_17","question":"A cube has how many edges?","options":["12", "8", "6", "16"],"correct":0,"marks":1,"topic":"Cube Folding","section":"Reasoning Ability"},
    {"id":"ra_fig_18","question":"Paper folded in half both ways and one corner cut. Unfolded holes = ?","options":["4", "1", "2", "8"],"correct":0,"marks":1,"topic":"Paper Cuts","section":"Reasoning Ability"},
    {"id":"ra_fig_19","question":"A cube with top=5, front=1, right=2. What is bottom?","options":["2", "4", "6", "3"],"correct":0,"marks":1,"topic":"Cube Folding","section":"Reasoning Ability"},
    {"id":"ra_fig_20","question":"Half circle cut from folded square. When unfolded forms a:","options":["Circle", "Semicircle", "Oval", "Full circle"],"correct":0,"marks":1,"topic":"Paper Cuts","section":"Reasoning Ability"},
    {"id":"ra_dm_1","question":"Statement: All cats are dogs. All dogs are birds. Conclusion: All cats are birds.","options":["True", "False", "Uncertain", "Partially true"],"correct":0,"marks":1,"topic":"Syllogism","section":"Reasoning Ability"},
    {"id":"ra_dm_2","question":"Statement: All students pass. Ram is a student. Conclusion: Ram passes.","options":["Definitely true", "Definitely false", "Uncertain", "Cannot say"],"correct":0,"marks":1,"topic":"Statement and Conclusion","section":"Reasoning Ability"},
    {"id":"ra_dm_3","question":"If all A are B and some B are C, then:","options":["All A are C", "Some A are C", "No A are C", "Cannot determine"],"correct":3,"marks":1,"topic":"Syllogism","section":"Reasoning Ability"},
    {"id":"ra_dm_4","question":"Data: X>Y, Y>Z. Is X>Z? Data sufficient?","options":["Yes", "No", "Partially", "Cannot say"],"correct":0,"marks":1,"topic":"Data Sufficiency","section":"Reasoning Ability"},
    {"id":"ra_dm_5","question":"Statement: No pen is a book. Some books are copies. Conclusion: Some copies are not pens.","options":["True", "False", "Uncertain", "Both true"],"correct":0,"marks":1,"topic":"Syllogism","section":"Reasoning Ability"},
    {"id":"ra_dm_6","question":"Premise: If it rains, ground is wet. Ground is wet. Conclusion: It rained.","options":["Valid", "Invalid", "Uncertain", "Correct"],"correct":1,"marks":1,"topic":"Statement and Conclusion","section":"Reasoning Ability"},
    {"id":"ra_dm_7","question":"All M are N. No N are P. Conclusion: No M are P.","options":["True", "False", "Uncertain", "Cannot say"],"correct":0,"marks":1,"topic":"Syllogism","section":"Reasoning Ability"},
    {"id":"ra_dm_8","question":"To find if a number is prime, need I: number itself, II: check up to \u221an. Both sufficient?","options":["Yes", "No", "Only I", "Only II"],"correct":3,"marks":1,"topic":"Data Sufficiency","section":"Reasoning Ability"},
    {"id":"ra_dm_9","question":"Some flowers are red. All red things are beautiful. Conclusion: Some flowers are beautiful.","options":["True", "False", "Partial", "Not determined"],"correct":0,"marks":1,"topic":"Statement and Conclusion","section":"Reasoning Ability"},
    {"id":"ra_dm_10","question":"All A are B. All B are C. No C are D. Conclusion: No A are D.","options":["Definitely true", "Definitely false", "Uncertain", "Not follow"],"correct":0,"marks":1,"topic":"Syllogism","section":"Reasoning Ability"},
    {"id":"ra_dm_11","question":"Statement: Every honest person is respected. He is not respected. Conclusion: He is not honest.","options":["Valid", "Invalid", "Uncertain", "Impossible"],"correct":0,"marks":1,"topic":"Statement and Conclusion","section":"Reasoning Ability"},
    {"id":"ra_dm_12","question":"Some boys play cricket. All cricket players are fit. Conclusion: Some boys are fit.","options":["Follows", "Does not follow", "Uncertain", "Cannot say"],"correct":0,"marks":1,"topic":"Syllogism","section":"Reasoning Ability"},
    {"id":"ra_dm_13","question":"I: a\u00b2+b\u00b2=c\u00b2 II: a=3, b=4. Is c=5? Sufficient from:","options":["I only", "II only", "Both needed", "Neither"],"correct":2,"marks":1,"topic":"Data Sufficiency","section":"Reasoning Ability"},
    {"id":"ra_dm_14","question":"No teacher is uneducated. All doctors are teachers. Conclusion: All doctors are educated.","options":["Follows", "Does not follow", "Uncertain", "Partially"],"correct":0,"marks":1,"topic":"Syllogism","section":"Reasoning Ability"},
    {"id":"ra_dm_15","question":"Premise: All mangoes are fruits. Premise: Some fruits are sweet. Conclusion: Some mangoes are sweet.","options":["True", "False", "Cannot be determined", "Always true"],"correct":2,"marks":1,"topic":"Statement and Conclusion","section":"Reasoning Ability"},
    {"id":"ra_dm_16","question":"All birds can fly. Penguin is a bird. Conclusion: Penguin can fly.","options":["Logically valid", "Logically invalid", "Uncertain", "Partially valid"],"correct":0,"marks":1,"topic":"Syllogism","section":"Reasoning Ability"},
    {"id":"ra_dm_17","question":"I: P>Q II: Q>R. Is P>R? Need:","options":["I only", "II only", "Both I & II", "Neither"],"correct":2,"marks":1,"topic":"Data Sufficiency","section":"Reasoning Ability"},
    {"id":"ra_dm_18","question":"Some cars are bikes. No bike is a truck. Conclusion: Some cars are not trucks.","options":["Follows", "Does not follow", "Partial", "Cannot say"],"correct":0,"marks":1,"topic":"Syllogism","section":"Reasoning Ability"},
    {"id":"ra_dm_19","question":"If P then Q. Q is false. What about P?","options":["P is true", "P is false", "P may be true", "P is unknown"],"correct":1,"marks":1,"topic":"Statement and Conclusion","section":"Reasoning Ability"},
    {"id":"ra_dm_20","question":"All X are Y. Some Z are X. Conclusion: Some Z are Y.","options":["True", "False", "Cannot determine", "Only sometimes"],"correct":0,"marks":1,"topic":"Syllogism","section":"Reasoning Ability"},
    {"id":"ra_vd_1","question":"In a Venn diagram, the intersection of two sets A and B represents:","options":["A∪B", "A∩B", "A only", "B only"],"correct":1,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_2","question":"All cats are animals. All animals have life. Where do cats fall in Venn diagram?","options":["Inside animals, inside life", "Outside both", "Inside animals only", "Inside life only"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_3","question":"Sets: Players, Singers, Actors. A person is all three. In Venn diagram they are in:","options":["All three intersections", "Two circles only", "One circle", "Outside all"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_4","question":"From Venn diagram: n(A)=30, n(B)=25, n(A\u2229B)=10. n(A\u222aB) = ?","options":["45", "55", "35", "65"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_5","question":"No circle overlaps another. This represents:","options":["Disjoint sets", "Subsets", "Equal sets", "Overlapping sets"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_6","question":"One circle completely inside another means:","options":["Subset", "Disjoint", "Equal", "Complement"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_7","question":"n(A)=20, n(A\u222aB)=30, n(A\u2229B)=5. n(B)=?","options":["15", "10", "20", "25"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_8","question":"From Venn DI: 60 like tea, 40 like coffee, 20 like both. Total who like at least one?","options":["80", "100", "60", "120"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_9","question":"Three overlapping circles represent Doctors, Engineers, MBA. Center region (all 3) represents:","options":["Professionals who are all three", "Only doctors", "None", "Doctors+Engineers"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_10","question":"n(A only) = n(A) \u2212 n(A\u2229B). If n(A)=15, n(A\u2229B)=5, n(A only)=?","options":["10", "20", "5", "15"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_11","question":"Venn shows: Animals-Fish-Whale. Whale is inside Fish, Fish inside Animals. Whale is:","options":["An animal and a fish", "Only a fish", "Only a whale", "Not an animal"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_12","question":"Universal set U=100, n(A)=60, n(A')=?","options":["40", "60", "100", "20"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_13","question":"Set A = multiples of 2, Set B = multiples of 3. A\u2229B = multiples of?","options":["6", "5", "2", "3"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_14","question":"In a survey 50 like A, 40 like B, 10 like both, 20 like neither. Total surveyed?","options":["100", "110", "90", "80"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_15","question":"Symmetric difference A\u25b3B means:","options":["Only in A or B, not both", "A∩B", "A∪B", "None of these"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_16","question":"From diagram: 5 only in A, 3 only in B, 2 in both. Total = ?","options":["10", "8", "12", "7"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_17","question":"A Venn diagram for 'All S are P' shows:","options":["S inside P", "S and P disjoint", "S=P", "P inside S"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_18","question":"n(A\u222aB) = n(A)+n(B)\u2212n(A\u2229B). If n(A)=8,n(B)=6,n(A\u2229B)=3. n(A\u222aB)=?","options":["11", "17", "14", "9"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_19","question":"In a class: 30 play cricket, 25 play football, 10 play both. Only cricket =","options":["20", "30", "15", "25"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vd_20","question":"Two equal overlapping circles A and B. Shaded region = A only. This is called:","options":["A-B", "A∩B", "A∪B", "A'"],"correct":0,"marks":1,"topic":"Logical Venn Diagram","section":"Reasoning Ability"},
    {"id":"ra_vis_1","question":"Mirror image of 'b' is:","options":["d", "q", "p", "b"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_2","question":"Water image of 'M' looks like:","options":["W", "Inverted M", "Rotated M", "N"],"correct":1,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_3","question":"A shape has 4 equal sides and 4 equal angles. It is a:","options":["Rhombus", "Rectangle", "Square", "Trapezium"],"correct":2,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_4","question":"Figure rotated 90\u00b0 clockwise: Arrow pointing Up becomes pointing:","options":["Right", "Left", "Down", "Up"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_5","question":"How many triangles in a star with 5 points (Star of David pattern)?","options":["5", "10", "6", "8"],"correct":1,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_6","question":"Mirror image: Number 6 is?","options":["9", "6", "b", "d"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_7","question":"If a figure has line symmetry, flipping it gives:","options":["Different figure", "Same figure", "Rotated figure", "Enlarged figure"],"correct":1,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_8","question":"How many squares in a 3\u00d73 grid?","options":["9", "14", "12", "16"],"correct":1,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_9","question":"Shape with all sides equal but no right angles is a:","options":["Square", "Rhombus", "Rectangle", "Parallelogram"],"correct":1,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_10","question":"A figure rotated 180\u00b0 looks like its:","options":["Mirror image", "Water image", "Same", "Enlarged form"],"correct":1,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_11","question":"Count triangles in figure with 4 equal triangles forming a larger triangle:","options":["4", "5", "6", "7"],"correct":2,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_12","question":"Water image is a reflection about which axis?","options":["Horizontal", "Vertical", "Diagonal", "Both"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_13","question":"A square and a triangle share one side. Total sides of combined figure?","options":["5", "6", "7", "4"],"correct":2,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_14","question":"How many rectangles in a 2\u00d73 grid?","options":["6", "12", "18", "9"],"correct":1,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_15","question":"Embedded figure: Which has a triangle hidden inside?","options":["Hexagon", "Pentagon", "All polygons can", "Square only"],"correct":2,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_16","question":"Figure turned upside down = rotated by:","options":["90°", "180°", "270°", "360°"],"correct":1,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_17","question":"Dots visible on top face of a cube when 1 is on bottom and 3 faces you: top is ?","options":["2", "4", "6", "5"],"correct":2,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_18","question":"Mirror image reverses:","options":["Top-bottom", "Left-right", "Both", "Neither"],"correct":1,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_19","question":"How many circles can be drawn through 3 non-collinear points?","options":["1", "Infinite", "0", "2"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"ra_vis_20","question":"Shape remaining after removing a square from a rectangle's corner?","options":["Pentagon", "Hexagon", "Triangle", "Trapezium"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Reasoning Ability"},
    {"id":"va_gr_1","question":"She __ (go) to school every day.","options":["go", "goes", "going", "went"],"correct":1,"marks":1,"topic":"Word Completion","section":"Verbal Ability"},
    {"id":"va_gr_2","question":"Complete: He is __ honest man.","options":["a", "an", "the", "no article"],"correct":1,"marks":1,"topic":"Sentence Completion","section":"Verbal Ability"},
    {"id":"va_gr_3","question":"Find error: 'She don't know the answer.'","options":["She", "don't", "know", "the answer"],"correct":1,"marks":1,"topic":"Error Identification","section":"Verbal Ability"},
    {"id":"va_gr_4","question":"Arrange: morning / in / he / jogs / the","options":["he jogs in the morning", "he in the morning jogs", "morning in he jogs the", "jogs he the in morning"],"correct":0,"marks":1,"topic":"Arrangement","section":"Verbal Ability"},
    {"id":"va_gr_5","question":"'Dear Sir' is used in __ letter.","options":["Formal", "Informal", "Personal", "Casual"],"correct":0,"marks":1,"topic":"Formal or Informal","section":"Verbal Ability"},
    {"id":"va_gr_6","question":"Join: 'She was tired. She kept working.'","options":["She was tired but she kept working", "She was tired so she kept working", "She was tired and she kept working", "She was tired or she kept working"],"correct":0,"marks":1,"topic":"Sentence Joining","section":"Verbal Ability"},
    {"id":"va_gr_7","question":"Meaning of 'Benevolent':","options":["Cruel", "Kind", "Ignorant", "Brave"],"correct":1,"marks":1,"topic":"Meanings","section":"Verbal Ability"},
    {"id":"va_gr_8","question":"Error: 'The team are playing well today.'","options":["The team", "are", "playing", "well today"],"correct":1,"marks":1,"topic":"Error Identification","section":"Verbal Ability"},
    {"id":"va_gr_9","question":"Complete: The project was completed __ schedule.","options":["in", "on", "at", "before"],"correct":1,"marks":1,"topic":"Sentence Completion","section":"Verbal Ability"},
    {"id":"va_gr_10","question":"'Yours faithfully' is used to close a __ letter.","options":["Formal", "Informal", "Love", "Birthday"],"correct":0,"marks":1,"topic":"Formal or Informal","section":"Verbal Ability"},
    {"id":"va_gr_11","question":"Synonym of 'Abundant':","options":["Scarce", "Plentiful", "Rare", "Common"],"correct":1,"marks":1,"topic":"Meanings","section":"Verbal Ability"},
    {"id":"va_gr_12","question":"Error: 'Neither he nor I are responsible.'","options":["Neither", "he", "are", "responsible"],"correct":2,"marks":1,"topic":"Error Identification","section":"Verbal Ability"},
    {"id":"va_gr_13","question":"Word completion: The __ of the meeting was productivity.","options":["goal", "agenda", "minutes", "chair"],"correct":1,"marks":1,"topic":"Word Completion","section":"Verbal Ability"},
    {"id":"va_gr_14","question":"Join: 'It was raining. We stayed inside.'","options":["It was raining so we stayed inside", "It was raining but we stayed inside", "It was raining and we stayed inside", "It was raining yet we stayed inside"],"correct":0,"marks":1,"topic":"Sentence Joining","section":"Verbal Ability"},
    {"id":"va_gr_15","question":"Antonym of 'Transparent':","options":["Clear", "Opaque", "Bright", "Visible"],"correct":1,"marks":1,"topic":"Meanings","section":"Verbal Ability"},
    {"id":"va_gr_16","question":"Complete: If I had money, I __ buy a car.","options":["will", "would", "should", "shall"],"correct":1,"marks":1,"topic":"Sentence Completion","section":"Verbal Ability"},
    {"id":"va_gr_17","question":"Correct arrangement: 'beautiful a is this painting'","options":["This is a beautiful painting", "A beautiful this painting is", "Beautiful is this a painting", "This painting beautiful a is"],"correct":0,"marks":1,"topic":"Arrangement","section":"Verbal Ability"},
    {"id":"va_gr_18","question":"Error: 'He suggested me to leave early.'","options":["He", "suggested", "me to", "leave early"],"correct":2,"marks":1,"topic":"Error Identification","section":"Verbal Ability"},
    {"id":"va_gr_19","question":"Meaning of 'Eloquent':","options":["Silent", "Well-spoken", "Rude", "Confused"],"correct":1,"marks":1,"topic":"Meanings","section":"Verbal Ability"},
    {"id":"va_gr_20","question":"'I miss you. I think of you daily.' \u2014 This is __ writing.","options":["Formal", "Informal", "Academic", "Professional"],"correct":1,"marks":1,"topic":"Formal or Informal","section":"Verbal Ability"},
    {"id":"va_rc_1","question":"RC: 'The Amazon rainforest produces 20% of world's oxygen.' Main idea?","options":["Rainforests are large", "Amazon is important for oxygen", "Trees produce oxygen", "World needs oxygen"],"correct":1,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_2","question":"Passage: 'Despite hardships, she persevered and succeeded.' Tone is:","options":["Pessimistic", "Optimistic", "Neutral", "Sarcastic"],"correct":1,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_3","question":"'Infer' means:","options":["State directly", "Draw conclusion", "Summarize", "Copy"],"correct":1,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_4","question":"RC: 'Global warming causes ice caps to melt, raising sea levels.' Effect of warming?","options":["More ice", "Rising seas", "Less water", "Cooler climate"],"correct":1,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_5","question":"Author uses 'Furthermore' to:","options":["Contradict", "Add information", "Conclude", "Summarize"],"correct":1,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_6","question":"'The protagonist was elated.' Elated means:","options":["Sad", "Angry", "Very happy", "Confused"],"correct":2,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_7","question":"RC purpose of 'However' in passage:","options":["To show contrast", "To add information", "To give example", "To conclude"],"correct":0,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_8","question":"Title best reflecting passage about water conservation:","options":["Save Water", "Water Pollution", "Water Cycle", "Rivers of India"],"correct":0,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_9","question":"'The study conclusively proved X.' 'Conclusively' means:","options":["Partially", "Definitely", "Approximately", "Tentatively"],"correct":1,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_10","question":"Passage: Technology improves life. But over-reliance is dangerous. Author's stance?","options":["Pro-technology", "Against technology", "Balanced view", "No opinion"],"correct":2,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_11","question":"RC: If author says 'contrary to popular belief', they are:","options":["Agreeing with common view", "Challenging common view", "Summarizing", "Giving examples"],"correct":1,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_12","question":"'Ambiguous' in a passage means the situation is:","options":["Clear", "Uncertain", "Definite", "Simple"],"correct":1,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_13","question":"RC: 'The policy was implemented despite opposition.' Despite implies:","options":["Because of", "Even though there was", "Without", "After"],"correct":1,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_14","question":"Passage about climate change ends 'Action is needed now.' Purpose is:","options":["Inform", "Persuade", "Entertain", "Describe"],"correct":1,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_15","question":"'The expert opined that...' Opined means:","options":["Demanded", "Stated as opinion", "Proved", "Announced"],"correct":1,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_16","question":"RC: Finding the main idea means finding:","options":["First sentence", "Last sentence", "Central theme", "Author's name"],"correct":2,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_17","question":"Passage: economy grew 7%. This is an example of:","options":["Inference", "Fact", "Opinion", "Assumption"],"correct":1,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_18","question":"'The results were unprecedented.' Unprecedented means:","options":["Expected", "Never happened before", "Ordinary", "Repeated"],"correct":1,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_19","question":"RC: Author describes problem then solution. Structure is:","options":["Compare-contrast", "Problem-solution", "Cause-effect", "Narrative"],"correct":1,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"va_rc_20","question":"Vocabulary: 'Diligent worker' means:","options":["Lazy", "Careless", "Hardworking", "Slow"],"correct":2,"marks":1,"topic":"Reading Comprehension","section":"Verbal Ability"},
    {"id":"aq_av_1","question":"Average of 10 numbers is 25. If one number 50 is added, new average?","options":["25", "26.19", "27", "30"],"correct":1,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_2","question":"Average of A and B is 20, B and C is 25, C and A is 15. A+B+C?","options":["60", "120", "90", "30"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_3","question":"Average temp Mon-Fri = 30\u00b0. Mon-Thu = 29\u00b0. Friday temp?","options":["34°", "35°", "33°", "32°"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_4","question":"Average of 5 consecutive odd numbers is 21. Largest?","options":["25", "27", "23", "29"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_5","question":"Average of first 20 natural numbers:","options":["10", "10.5", "11", "9.5"],"correct":1,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_6","question":"Average runs in 15 matches = 40. Need 60 average in 20 matches. Total runs in next 5?","options":["500", "400", "300", "600"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_7","question":"Weighted average: 30 students avg 60, 20 students avg 80. Overall avg?","options":["68", "70", "72", "65"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_8","question":"Average of p, q, r = 10. Average of p+q, q+r, r+p = ?","options":["15", "20", "10", "30"],"correct":1,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_9","question":"If average of x, x+2, x+4, x+6 = 20, then x = ?","options":["17", "18", "16", "15"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_10","question":"Class of 40: avg marks = 75. If top 10 avg = 85, bottom 30 avg = ?","options":["71.67", "70", "72", "73"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_11","question":"Average of squares of 1, 2, 3, 4, 5:","options":["11", "9", "10", "12"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_12","question":"If mean of 100 items = 50, and 2 items 40 and 60 are removed, new mean?","options":["50", "48", "52", "51"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_13","question":"Bowling average: 40 wickets in 900 runs. Avg runs per wicket?","options":["22.5", "20", "25", "18"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_14","question":"Sum of 5 numbers is 265. Average of first two is 43, last two is 59. Middle?","options":["45", "55", "65", "60"],"correct":1,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_15","question":"Average age of 3 friends is 24. If one is 30, other two average?","options":["21", "22", "20", "18"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_16","question":"Average = 40. After removing an observation the new avg = 39.5 for n-1 items. What was removed from n=20?","options":["50", "60", "49", "55"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_17","question":"Mean of 50, 60, 70, 80, 90 = ?","options":["70", "75", "65", "68"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_18","question":"Average of 6 numbers = 12. Sum = ?","options":["72", "60", "84", "48"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_19","question":"A man's avg expenditure for 4 months is \u20b92700, next 8 months \u20b92900. Annual avg?","options":["₹2833", "₹2800", "₹2900", "₹2750"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_av_20","question":"Average weight of 8 persons increases by 2.5 when one 60 kg leaves and new person joins. New person's weight?","options":["80", "75", "85", "70"],"correct":0,"marks":1,"topic":"Advanced Averages","section":"Advanced Quant"},
    {"id":"aq_pl_1","question":"CP=\u20b9500, SP=\u20b9600. Profit%?","options":["15%", "20%", "25%", "10%"],"correct":1,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_2","question":"Loss%=10%, SP=\u20b9450. CP=?","options":["₹500", "₹400", "₹550", "₹480"],"correct":0,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_3","question":"Successive discounts 20% and 10% on \u20b91000. Final SP?","options":["₹720", "₹700", "₹730", "₹750"],"correct":0,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_4","question":"Marked price \u20b9800, discount 15%, SP?","options":["₹680", "₹660", "₹640", "₹700"],"correct":0,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_5","question":"Profit on selling 20 articles = CP of 5 articles. Profit%?","options":["20%", "25%", "33.3%", "30%"],"correct":1,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_6","question":"CP=\u20b91200, profit=25%. SP=?","options":["₹1500", "₹1400", "₹1600", "₹1450"],"correct":0,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_7","question":"Two articles each sold at \u20b9900, one with 20% profit, one with 20% loss. Overall?","options":["No loss no profit", "₹75 loss", "₹75 profit", "₹50 loss"],"correct":1,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_8","question":"A merchant uses 900g weight instead of 1kg. Profit%?","options":["10%", "11.11%", "9.09%", "12%"],"correct":1,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_9","question":"To gain 25% after giving 15% discount, mark price above CP by?","options":["47.06%", "40%", "50%", "35%"],"correct":0,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_10","question":"If SP = 4/3 of CP, profit%?","options":["25%", "33.3%", "30%", "20%"],"correct":1,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_11","question":"CP=\u20b92400, overhead=\u20b9600, profit wanted=20%. SP=?","options":["₹3600", "₹3000", "₹3200", "₹2880"],"correct":0,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_12","question":"Book sold at 5% loss. Had it sold for \u20b924 more, gain would be 7%. CP=?","options":["₹200", "₹175", "₹180", "₹220"],"correct":0,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_13","question":"If goods bought at \u20b9100 and sold at \u20b980, loss%?","options":["20%", "25%", "15%", "18%"],"correct":0,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_14","question":"Selling 100 oranges at profit of 2 oranges. Profit%?","options":["2%", "1.96%", "2.04%", "2.5%"],"correct":2,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_15","question":"Trade discount 30%, cash discount 10% on \u20b91000. Net price?","options":["₹630", "₹600", "₹700", "₹650"],"correct":0,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_16","question":"Ratio of CP:SP = 5:6. Profit%?","options":["20%", "16.67%", "25%", "15%"],"correct":0,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_17","question":"Gain on 50 items = loss on 40 items. Ratio of SP to CP?","options":["9:8 (when selling 50)", "1:1", "90:80", "8:9"],"correct":0,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_18","question":"A sells to B at 20% profit. B sells to C at 10% loss. C pays \u20b91080. A's CP?","options":["₹1000", "₹1100", "₹900", "₹1200"],"correct":0,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_19","question":"CP of 20 items = SP of 15 items. Profit or loss%?","options":["25% profit", "25% loss", "33.3% profit", "33.3% loss"],"correct":2,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_pl_20","question":"False weight shows 800g as 1kg. Profit% if selling at CP?","options":["20%", "25%", "16.67%", "22.5%"],"correct":1,"marks":1,"topic":"Profit Loss","section":"Advanced Quant"},
    {"id":"aq_hl_1","question":"LCM of 12, 15, 20 is:","options":["60", "30", "120", "90"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_2","question":"HCF of 36, 48, 60 is:","options":["6", "12", "18", "24"],"correct":1,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_3","question":"Product of two numbers = 1920, HCF = 16. LCM = ?","options":["120", "60", "240", "180"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_4","question":"LCM of two co-prime numbers 8 and 15 is:","options":["120", "60", "240", "30"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_5","question":"Find the greatest number that divides 25, 45, 65 leaving remainder 5.","options":["10", "20", "5", "15"],"correct":1,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_6","question":"Least number divisible by 2,3,4,5,6 is:","options":["60", "30", "120", "180"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_7","question":"HCF of two numbers is 12, LCM is 180. One is 36, other is?","options":["60", "48", "72", "84"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_8","question":"Three bells ring at 15, 20, 30 min intervals. All ring together at 6am, next at?","options":["7am", "8am", "9am", "10am"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_9","question":"HCF of (a\u00b2\u2212b\u00b2) and (a\u2212b) is:","options":["a−b", "a+b", "a²−b²", "ab"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_10","question":"Greatest 4-digit number divisible by 12, 15, 18:","options":["9720", "9900", "9990", "9720"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_11","question":"HCF of 0.36 and 0.48 is:","options":["0.12", "0.06", "0.18", "0.24"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_12","question":"LCM of (1/2),(2/3),(3/4) is:","options":["3", "6", "12", "1"],"correct":1,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_13","question":"Smallest number which when divided by 4,6,8 leaves remainder 2?","options":["26", "22", "14", "18"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_14","question":"HCF \u00d7 LCM of two numbers = product of the two numbers. This is:","options":["Always true", "Never true", "Sometimes true", "True only for primes"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_15","question":"Find LCM of 2\u2075 \u00d7 3\u00b2 and 2\u00b3 \u00d7 3\u2074:","options":["2⁵ × 3⁴", "2³ × 3²", "2⁸ × 3⁶", "2⁵ × 3²"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_16","question":"Number of pairs of integers whose HCF=16 and LCM=240:","options":["2", "3", "4", "5"],"correct":1,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_17","question":"Greatest number dividing 4053 and 12909 leaving remainder 3?","options":["25", "50", "75", "100"],"correct":1,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_18","question":"LCM of first 10 natural numbers:","options":["2520", "5040", "1260", "720"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_19","question":"If HCF of a,b = 1 (coprime), LCM of a,b = ?","options":["a×b", "a+b", "a−b", "a/b"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_hl_20","question":"Three runners complete a lap in 6, 8, 12 min. Meet at start after?","options":["24 min", "48 min", "12 min", "36 min"],"correct":0,"marks":1,"topic":"HCF LCM","section":"Advanced Quant"},
    {"id":"aq_pc_1","question":"\u2075P\u2083 = ?","options":["60", "120", "20", "30"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_2","question":"\u2075C\u2082 = ?","options":["10", "20", "15", "5"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_3","question":"Number of ways to arrange letters of EXAM?","options":["24", "12", "48", "6"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_4","question":"4 boys and 3 girls. Committee of 3: 1 girl and 2 boys. Ways?","options":["18", "12", "24", "36"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_5","question":"Number of 3-digit numbers from digits 1-5 without repetition?","options":["60", "120", "24", "48"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_6","question":"In how many ways can 5 books be arranged on a shelf?","options":["120", "60", "24", "720"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_7","question":"From 10 questions, answer any 7. Ways?","options":["120", "210", "90", "180"],"correct":1,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_8","question":"Round table arrangement of 5 people = ?","options":["24", "120", "60", "48"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_9","question":"Word MATHEMATICS has how many arrangements?","options":["4989600", "39916800", "2494800", "9979200"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_10","question":"Number of diagonals in a hexagon?","options":["9", "12", "6", "15"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_11","question":"nCr = nC(n-r). \u2078C\u2083 = ?","options":["56", "28", "70", "42"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_12","question":"How many ways to choose 3 from 8 candidates?","options":["56", "28", "168", "336"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_13","question":"LEVEL \u2014 arrangements with all L's together?","options":["12", "24", "6", "18"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_14","question":"Password: 3 digits, 2 letters (26 letters, 10 digits), no repetition. Count?","options":["1404000", "702000", "2808000", "468000"],"correct":1,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_15","question":"Team of 11 from 15: 5 must be included. Ways?","options":["210", "126", "252", "168"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_16","question":"n! / (n-3)! = 504. n = ?","options":["9", "8", "7", "10"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_17","question":"Number of triangles from 6 points on circle?","options":["20", "15", "10", "25"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_18","question":"Ways to distribute 5 distinct books among 3 students?","options":["243", "125", "15", "60"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_19","question":"Sum of all 4-digit numbers using 1,2,3,4 (no repeat)?","options":["66660", "55550", "44440", "33330"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pc_20","question":"Handshakes among 10 people (each shakes hand with every other once)?","options":["45", "90", "100", "50"],"correct":0,"marks":1,"topic":"Permutation and Combination","section":"Advanced Quant"},
    {"id":"aq_pr_1","question":"P(A) = 0.6, P(B) = 0.4, P(A\u2229B) = 0.2. P(A\u222aB)?","options":["0.8", "0.6", "1.0", "0.4"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_2","question":"Probability of getting head in a coin toss:","options":["1/2", "1/4", "1/3", "2/3"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_3","question":"Probability of sum=7 when two dice thrown?","options":["1/6", "7/36", "6/36", "1/4"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_4","question":"Bag: 3 red, 4 blue. P(red drawn)?","options":["3/7", "4/7", "1/2", "3/4"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_5","question":"P(A) = 1/3, P(B) = 1/4, independent. P(A and B)?","options":["1/12", "1/7", "7/12", "1/6"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_6","question":"Two cards drawn from deck of 52. P(both aces)?","options":["1/221", "1/52", "4/52", "1/169"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_7","question":"P(A) = 0.7. P(not A) = ?","options":["0.3", "0.7", "1.0", "0.4"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_8","question":"10 students, 3 selected. P(specific 2 are selected)?","options":["3/10", "1/15", "1/10", "3/15"],"correct":1,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_9","question":"Probability of getting at least one tail in 3 coin tosses?","options":["7/8", "1/8", "3/4", "5/8"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_10","question":"A bag has 5W, 3B balls. Two drawn with replacement. P(both white)?","options":["25/64", "15/64", "10/64", "20/64"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_11","question":"Expected value of a fair die throw?","options":["3.5", "3", "4", "2.5"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_12","question":"P(Ace from a deck) = ?","options":["1/13", "4/52 = 1/13", "1/4", "Both A and B"],"correct":3,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_13","question":"In Binomial distribution, P(X=x) = nCx p^x q^(n-x). If n=3, p=0.5, P(X=2)?","options":["3/8", "1/8", "6/8", "2/8"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_14","question":"Mutually exclusive events: P(A or B) = ?","options":["P(A) + P(B)", "P(A) × P(B)", "P(A) + P(B) - P(A∩B)", "P(A|B)"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_15","question":"P(A)=0.4, P(B|A)=0.5. P(A\u2229B)=?","options":["0.2", "0.4", "0.5", "0.9"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_16","question":"Probability of NOT getting 6 on a single die roll?","options":["5/6", "1/6", "4/6", "2/3"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_17","question":"In a group of 3, P(at least 2 boys) if P(boy)=0.5?","options":["1/2", "3/8", "1/4", "5/8"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_18","question":"Two events are independent if P(A\u2229B) = ?","options":["P(A)+P(B)", "P(A)×P(B)", "P(A)-P(B)", "1"],"correct":1,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_19","question":"A card is drawn from 52. P(king or red)?","options":["7/13", "28/52", "15/26", "1/2"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pr_20","question":"In a lottery of 100 tickets, 5 win. P(winning with 3 tickets)?","options":["3/20", "3/100", "5/100", "1/20"],"correct":0,"marks":1,"topic":"Probability","section":"Advanced Quant"},
    {"id":"aq_pz_1","question":"A farmer has 17 sheep, all but 9 die. How many left?","options":["9", "8", "17", "0"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_2","question":"How many months have 28 days?","options":["1", "2", "12", "All months"],"correct":3,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_3","question":"A clock shows 3:15. Angle between hands?","options":["7.5°", "0°", "15°", "22.5°"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_4","question":"Divide 30 by half and add 10. Answer?","options":["25", "70", "20", "15"],"correct":1,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_5","question":"If 5+3=28, 9+1=810, then 8+6=?","options":["214", "48", "1448", "1414"],"correct":3,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_6","question":"Missing number: 2, 3, 5, 8, 13, ?","options":["21", "20", "18", "24"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_7","question":"3 cats catch 3 mice in 3 minutes. 100 cats catch 100 mice in?","options":["3 min", "100 min", "33 min", "1 min"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_8","question":"Age puzzle: 5 years ago A was 5\u00d7 B's age. Now 3\u00d7 B. A's present age?","options":["30", "25", "35", "20"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_9","question":"Rs 50 note is forged. Shopkeeper accepts it, gives \u20b940 change and goods worth \u20b910. Loss?","options":["₹50", "₹40", "₹100", "₹90"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_10","question":"1/3 of 12 + half of 14 = ?","options":["11", "13", "10", "9"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_11","question":"Word puzzle: What has cities but no houses, mountains but no trees, water but no fish?","options":["A map", "A globe", "A picture", "A book"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_12","question":"How many times digit 1 appears between 1 and 100?","options":["21", "11", "20", "22"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_13","question":"Number that equals 5 times the sum of its digits?","options":["10", "45", "0", "5"],"correct":1,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_14","question":"ABCDE \u00d7 4 = EDCBA. A=?","options":["8", "2", "9", "1"],"correct":2,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_15","question":"If a box has 2 red, 3 blue balls drawn blindly. Minimum draws to guarantee 2 of same color?","options":["3", "4", "2", "5"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_16","question":"Train 1 km long, speed 60 km/h crosses tunnel 1 km long. Time?","options":["2 min", "1 min", "3 min", "1.5 min"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_17","question":"A brick weighs 1 kg + half a brick. Full brick weighs?","options":["2 kg", "1.5 kg", "3 kg", "2.5 kg"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_18","question":"A is 2 years older than B. In 10 years, A is 1.5\u00d7 B's age now + 2. A's age now?","options":["12", "14", "10", "8"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_19","question":"If 1 man digs a hole in 3 days, 3 men dig in?","options":["1 day", "3 days", "9 days", "2 days"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"aq_pz_20","question":"Two frogs in a well, 10m deep. Each jump 2m, fall 1m. Reach top in?","options":["9 jumps", "10 jumps", "8 jumps", "5 jumps"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Quant"},
    {"id":"ar_da_1","question":"6 books A-F. A before B. C not adjacent to B. D last. E before F. First book?","options":["A", "C", "E", "F"],"correct":2,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_2","question":"5 friends sit in row: P, Q, R, S, T. P not at ends, Q next to P. T at right end. Who is at left end?","options":["R", "S", "Q", "T"],"correct":1,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_3","question":"Rank: A>B, C>D, B>D, A<C. Second highest?","options":["A", "B", "C", "D"],"correct":0,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_4","question":"7 boxes stacked: Red on Blue, Green above Yellow, Blue not at bottom, Yellow at bottom. Top box?","options":["Red", "Blue", "Green", "Yellow"],"correct":2,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_5","question":"Students in row: Taller than 3 = 4th from right. Shorter than 5 = 6th from left. Total?","options":["12", "10", "11", "9"],"correct":1,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_6","question":"5 items A,B,C,D,E. A costs more than B but less than C. D costs most. E costs least. 2nd costliest?","options":["C", "A", "B", "E"],"correct":0,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_7","question":"Schedule: Mon-Fri. A on Mon. B not on Fri. C after B. D before A. E on Fri. D on?","options":["Mon", "Sun (before Mon)", "Not possible", "Sat"],"correct":1,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_8","question":"6 floors 1-6, bottom=1. A on 3, B above A, C below A, D on top, E between A and B. E on floor?","options":["4", "5", "2", "3"],"correct":0,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_9","question":"Cricket batting order: 1-11. M bats 3rd. N bats after L. K bats before M. L 5th. N bats?","options":["6th", "7th", "4th", "8th"],"correct":0,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_10","question":"Ages: P>Q, R<S, S<P, Q>R. Youngest?","options":["R", "Q", "P", "S"],"correct":0,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_11","question":"A,B,C,D,E in a row. C in middle. A to left of C. B to right of D. E at right end. A's position?","options":["1st", "2nd", "3rd", "4th"],"correct":1,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_12","question":"8 players tournament. Round robin, each plays each once. Total matches?","options":["28", "56", "32", "24"],"correct":0,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_13","question":"5 cities PQRST on a line. Q between P and R. T rightmost. S between R and T. Leftmost?","options":["P", "Q", "R", "S"],"correct":0,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_14","question":"4 persons A,B,C,D born on different days of week. A on Wednesday. B 2 days after A. C day before B. D?","options":["Sunday", "Monday", "Tuesday", "Saturday"],"correct":3,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_15","question":"Books on shelf: Maths to left of Science. English rightmost. Hindi left of Maths. Order from left?","options":["Hindi,Maths,Science,English", "Maths,Hindi,English,Science", "English,Science,Maths,Hindi", "Science,Hindi,Maths,English"],"correct":0,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_16","question":"Alphabetical arrangement: CAT, BAT, RAT, MAT, PAT. 3rd when sorted?","options":["MAT", "PAT", "RAT", "CAT"],"correct":0,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_17","question":"Time arrangement: Events P,Q,R,S. S before Q. R after P. Q before P. Order?","options":["S,Q,P,R", "R,P,Q,S", "P,R,Q,S", "Q,P,S,R"],"correct":0,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_18","question":"Salary: A>B>C>D. E earns between B and C. Who earns 3rd highest?","options":["E", "B", "C", "D"],"correct":0,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_19","question":"Row: X is 5th from left, Y is 5th from right. 3 between them. Total?","options":["13", "12", "11", "14"],"correct":0,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_da_20","question":"Speed ranking: P fastest, Q slower than R, S faster than Q but slower than R. Order fastest first?","options":["P R S Q", "P S R Q", "P Q R S", "R P S Q"],"correct":0,"marks":1,"topic":"Data Arrangements","section":"Advanced Reasoning"},
    {"id":"ar_vr_1","question":"Figure: Triangle inside square inside circle. How many shapes?","options":["3", "6", "4", "5"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_2","question":"Pattern: \u25a1 \u25b3 \u25a1 \u25b3 \u25a1 _ What comes next?","options":["□", "△", "○", "◇"],"correct":1,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_3","question":"Matrix: 1st row 1,2,3; 2nd row 4,5,6; 3rd row 7,8,?","options":["9", "10", "11", "12"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_4","question":"Left half of symmetric figure shown. Right half mirrors it. Shape formed?","options":["Same shape", "Different shape", "Larger shape", "Inverted shape"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_5","question":"3D figure shows cube with one corner removed. Remaining vertices?","options":["7", "6", "8", "5"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_6","question":"Series: Big circle \u2192 Small circle \u2192 Triangle \u2192 Big circle \u2192 Small circle \u2192 ?","options":["Triangle", "Square", "Big circle", "Diamond"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_7","question":"Shadow of a sphere is a:","options":["Circle", "Sphere", "Ellipse", "Square"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_8","question":"Analogy: Big:Small :: Dark:?","options":["Bright", "Light", "Night", "Color"],"correct":1,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_9","question":"Rotating a square 45\u00b0 gives:","options":["Diamond shape", "Same square", "Rectangle", "Triangle"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_10","question":"Number of lines of symmetry in a regular pentagon?","options":["5", "4", "3", "10"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_11","question":"Pattern: \u2605 \u2605\u2605 \u2605\u2605\u2605 \u2605\u2605\u2605\u2605 \u2014 next?","options":["★★★★★", "★★★★", "★★★★★★", "★★★"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_12","question":"Odd figure out: square, rectangle, rhombus, triangle, parallelogram","options":["Square", "Rectangle", "Triangle", "Parallelogram"],"correct":2,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_13","question":"Two overlapping circles create how many distinct regions?","options":["3", "2", "4", "5"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_14","question":"Figure matrix: row adds up; column 1=2,4,6. Next number in column?","options":["8", "10", "7", "9"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_15","question":"A pentagon has how many diagonals?","options":["5", "3", "7", "4"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_16","question":"If the hour hand points to 3 and minute hand to 12, angle = ?","options":["90°", "60°", "120°", "45°"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_17","question":"Embedded shape: A hexagon contains how many triangles (from centre)?","options":["6", "4", "3", "8"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_18","question":"Shape with exactly one line of symmetry?","options":["Isosceles triangle", "Equilateral triangle", "Square", "Circle"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_19","question":"3 straight cuts on a circle \u2014 maximum regions created?","options":["7", "6", "4", "5"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_vr_20","question":"Mirror image of time 06:25 as seen in a clock mirror?","options":["05:35", "06:35", "05:35", "07:35"],"correct":0,"marks":1,"topic":"Visual Reasoning","section":"Advanced Reasoning"},
    {"id":"ar_pz_1","question":"River crossing: Farmer, Fox, Chicken, Grain. Minimum trips?","options":["7", "5", "9", "6"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_2","question":"Weighing puzzle: 9 balls, 1 heavier. Minimum weighings to find?","options":["2", "3", "1", "4"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_3","question":"Liar/Truth-teller: A says 'B is a liar'. B says 'A and I are same'. B is?","options":["Truth-teller", "Liar", "Cannot determine", "Neither"],"correct":1,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_4","question":"Grid puzzle: Knight's move on 4\u00d74 board from A1 to D4, minimum moves?","options":["4", "3", "5", "6"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_5","question":"12 coins, 1 counterfeit (lighter or heavier). Minimum balance weighings?","options":["3", "4", "2", "5"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_6","question":"Two ropes each burn in 60 minutes non-uniformly. Measure 45 minutes?","options":["Light both ends of one, one end of other", "Impossible", "Use timing", "Light one rope"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_7","question":"Tower of Hanoi with 3 disks. Minimum moves?","options":["7", "8", "6", "9"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_8","question":"Fill 4L jug using 3L and 5L jugs:","options":["Possible", "Not possible", "Need 4L exactly", "Two steps"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_9","question":"100 lockers, 100 students. After all rounds, how many open?","options":["10", "100", "1", "50"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_10","question":"Three missionaries, three cannibals cross river. Boat holds 2. Safe crossings?","options":["Possible", "Impossible", "Only 2 trips", "4 trips"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_11","question":"Pigeon hole: 367 people \u2014 at least 2 share same birthday?","options":["Yes always", "Not guaranteed", "Maybe", "Only in leap year"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_12","question":"Handshakes in group of 10 where each shakes hand with 3 others exactly?","options":["15", "30", "12", "20"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_13","question":"Magic square 3\u00d73 using 1-9, sum per row=15. Centre number?","options":["5", "4", "6", "3"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_14","question":"Hat puzzle: 3 hats from 2 black 3 white. Person at back sees 2 white says 'I know mine'. Colour?","options":["Black", "White", "Cannot know", "Either"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_15","question":"Sequence: 1 rabbit sees 6 elephants each with 2 monkeys, each monkey has 1 parrot. How many animals going to river?","options":["1", "25", "10", "15"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_16","question":"2 trains approach each other, 100 km apart, 50 km/h each. Fly at 200 km/h bounces between. Distance flown?","options":["200 km", "100 km", "150 km", "400 km"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_17","question":"Prisoner chooses door: 2 doors, 1 car 1 goat. Host reveals goat behind another. Switch?","options":["Yes, doubles chance", "No, same chance", "Never switch", "Host choice matters"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_18","question":"Graph: 7 bridges of K\u00f6nigsberg. Euler path exists if vertices with odd degree \u2264","options":["2", "4", "0", "Any"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_19","question":"Sand timer: 4 min + 7 min timer. Measure 9 minutes?","options":["Possible", "Impossible", "Need 9 min", "Only 11 min"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"},
    {"id":"ar_pz_20","question":"Coin puzzle: 6 coins, 3 heads 3 tails, blindfolded. Split into 2 groups of 3, equal heads each?","options":["Always possible", "Impossible", "Flip 3 coins", "Random"],"correct":0,"marks":1,"topic":"Puzzles","section":"Advanced Reasoning"}
]

# ── Part B: SQL Questions ──────────────────────────────────────────────────────
SQL_BANK = [
    {
        "id":"s1","marks":5,"topic":"SELECT",
        "question":"Write a SQL query to find all employees with salary greater than 50000.",
        "setup_sql": """CREATE TABLE employees (
    id INTEGER PRIMARY KEY,
    name TEXT,
    department TEXT,
    salary REAL
);
INSERT INTO employees VALUES
(1,'Alice','Engineering',75000),
(2,'Bob','Marketing',45000),
(3,'Carol','Engineering',90000),
(4,'Dave','HR',40000),
(5,'Eve','Engineering',55000);""",
        "expected_output": [["Alice","Engineering",75000.0],["Carol","Engineering",90000.0],["Eve","Engineering",55000.0]],
        "expected_cols": ["name","department","salary"],
        "hint": "Use WHERE salary > 50000"
    },
    {
        "id":"s2","marks":5,"topic":"GROUP BY",
        "question":"Write a SQL query to count the number of employees in each department.",
        "setup_sql": """CREATE TABLE employees (
    id INTEGER PRIMARY KEY,
    name TEXT,
    department TEXT,
    salary REAL
);
INSERT INTO employees VALUES
(1,'Alice','Engineering',75000),
(2,'Bob','Marketing',45000),
(3,'Carol','Engineering',90000),
(4,'Dave','HR',40000),
(5,'Eve','Engineering',55000);""",
        "expected_output": [["Engineering",3],["HR",1],["Marketing",1]],
        "expected_cols": ["department","count"],
        "hint": "Use GROUP BY with COUNT(*)"
    },
    {
        "id":"s3","marks":5,"topic":"JOIN",
        "question":"Write a SQL query to get each employee's name and their manager's name using a self-join.",
        "setup_sql": """CREATE TABLE employees (
    id INTEGER PRIMARY KEY,
    name TEXT,
    manager_id INTEGER
);
INSERT INTO employees VALUES
(1,'Alice',NULL),
(2,'Bob',1),
(3,'Carol',1),
(4,'Dave',2);""",
        "expected_output": [["Bob","Alice"],["Carol","Alice"],["Dave","Bob"]],
        "expected_cols": ["employee","manager"],
        "hint": "Self-join employees table on manager_id = id"
    },
    {
        "id":"s4","marks":5,"topic":"Aggregation",
        "question":"Write a SQL query to find the highest salary in each department.",
        "setup_sql": """CREATE TABLE employees (
    id INTEGER PRIMARY KEY,
    name TEXT,
    department TEXT,
    salary REAL
);
INSERT INTO employees VALUES
(1,'Alice','Engineering',75000),
(2,'Bob','Marketing',45000),
(3,'Carol','Engineering',90000),
(4,'Dave','HR',40000),
(5,'Eve','Engineering',55000);""",
        "expected_output": [["Engineering",90000.0],["HR",40000.0],["Marketing",45000.0]],
        "expected_cols": ["department","max_salary"],
        "hint": "Use MAX() with GROUP BY department"
    },
    {
        "id":"s5","marks":5,"topic":"Subquery",
        "question":"Write a SQL query to find employees earning above the average salary.",
        "setup_sql": """CREATE TABLE employees (
    id INTEGER PRIMARY KEY,
    name TEXT,
    salary REAL
);
INSERT INTO employees VALUES
(1,'Alice',75000),(2,'Bob',45000),(3,'Carol',90000),
(4,'Dave',40000),(5,'Eve',55000);""",
        "expected_output": [["Alice",75000.0],["Carol",90000.0]],
        "expected_cols": ["name","salary"],
        "hint": "Use WHERE salary > (SELECT AVG(salary) FROM employees)"
    },
]

# ── Part C: Coding Questions ───────────────────────────────────────────────────
CODING_BANK = [
    {"id":"c_twosum","title":"Two Sum","marks":10,"topic":"Arrays","question":"Given an array of integers nums and an integer target, return the indices of the two numbers that add up to target. You may assume that each input has exactly one solution, and you may not use the same element twice.","hint":"Use a hash map to store each number and its index. For each number, check if (target - number) is already in the map.","test_cases":[{"input": "4\n2 7 11 15\n9", "expected_output": "0 1"}, {"input": "3\n3 2 4\n6", "expected_output": "1 2"}, {"input": "2\n3 3\n6", "expected_output": "0 1"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\ntarget = int(input())\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        int target = sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_removedup","title":"Remove Duplicates from Sorted Array","marks":10,"topic":"Arrays","question":"Given a sorted array nums, remove duplicates in-place such that each element appears only once. Return the length k of the array with unique elements. Print each unique element on a new line.","hint":"Use two pointers. One pointer tracks the position of the last unique element.","test_cases":[{"input": "5\n1 1 2 2 3", "expected_output": "3\n1\n2\n3"}, {"input": "7\n0 0 1 1 1 2 2", "expected_output": "3\n0\n1\n2"}, {"input": "3\n1 2 3", "expected_output": "3\n1\n2\n3"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_stockbest","title":"Best Time to Buy and Sell Stock","marks":10,"topic":"Arrays","question":"Given an array prices where prices[i] is the price of a stock on day i, return the maximum profit. You can only hold one stock at a time. If no profit possible, return 0.","hint":"Track the minimum price seen so far. At each step, compute profit = current - min_price.","test_cases":[{"input": "6\n7 1 5 3 6 4", "expected_output": "5"}, {"input": "5\n7 6 4 3 1", "expected_output": "0"}, {"input": "4\n1 2 3 4", "expected_output": "3"}],"template_python":"n = int(input())\nprices = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] prices = new int[n];\n        for(int i=0;i<n;i++) prices[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_plusone","title":"Plus One","marks":10,"topic":"Arrays","question":"Given a large integer represented as an array of digits (most significant digit first), increment it by one. Print the resulting digits separated by spaces.","hint":"Start from the last digit. Add 1. If digit becomes 10, set to 0 and carry 1 to next digit.","test_cases":[{"input": "3\n1 2 3", "expected_output": "1 2 4"}, {"input": "3\n9 9 9", "expected_output": "1 0 0 0"}, {"input": "1\n9", "expected_output": "1 0"}],"template_python":"n = int(input())\ndigits = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] d = new int[n];\n        for(int i=0;i<n;i++) d[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_missingnum","title":"Missing Number","marks":10,"topic":"Arrays","question":"Given an array nums containing n distinct numbers in the range [0, n], return the only number in the range that is missing.","hint":"Expected sum of 0..n is n*(n+1)/2. Subtract actual sum.","test_cases":[{"input": "3\n3 0 1", "expected_output": "2"}, {"input": "1\n0", "expected_output": "1"}, {"input": "4\n9 6 4 2 3 5 7 0 1", "expected_output": "8"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_maxsubarray","title":"Maximum Subarray","marks":10,"topic":"Arrays","question":"Given an integer array nums, find the subarray with the largest sum and return its sum. (Kadane's Algorithm)","hint":"Track current_sum and max_sum. At each element, current_sum = max(num, current_sum + num).","test_cases":[{"input": "9\n-2 1 -3 4 -1 2 1 -5 4", "expected_output": "6"}, {"input": "1\n1", "expected_output": "1"}, {"input": "5\n5 4 -1 7 8", "expected_output": "23"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_movezeroes","title":"Move Zeroes","marks":10,"topic":"Arrays","question":"Given an integer array nums, move all 0's to the end while maintaining relative order of non-zero elements. Print the result.","hint":"Use two pointers. One pointer places non-zero elements at the front.","test_cases":[{"input": "5\n0 1 0 3 12", "expected_output": "1 3 12 0 0"}, {"input": "2\n0 1", "expected_output": "1 0"}, {"input": "3\n0 0 1", "expected_output": "1 0 0"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_containsdup","title":"Contains Duplicate","marks":10,"topic":"Arrays","question":"Given an integer array nums, return true if any value appears at least twice, and false if every element is distinct.","hint":"Use a hash set. If element already in set, return true.","test_cases":[{"input": "4\n1 2 3 1", "expected_output": "true"}, {"input": "4\n1 2 3 4", "expected_output": "false"}, {"input": "5\n1 1 1 3 3", "expected_output": "true"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_intersect","title":"Intersection of Two Arrays II","marks":10,"topic":"Arrays","question":"Given two integer arrays nums1 and nums2, return an array of their intersection including duplicates. The result can be in any order.","hint":"Use a hash map to count occurrences in nums1. For each element in nums2, if count > 0, add to result.","test_cases":[{"input": "4\n1 2 2 1\n3\n2 2 3", "expected_output": "2 2"}, {"input": "4\n4 9 5 3\n4\n9 4 9 8", "expected_output": "4 9"}, {"input": "3\n1 2 3\n2\n4 5", "expected_output": ""}],"template_python":"n1 = int(input())\nnums1 = list(map(int, input().split()))\nn2 = int(input())\nnums2 = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n1=sc.nextInt(); int[]nums1=new int[n1]; for(int i=0;i<n1;i++)nums1[i]=sc.nextInt();\n        int n2=sc.nextInt(); int[]nums2=new int[n2]; for(int i=0;i<n2;i++)nums2[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_rotate","title":"Rotate Array","marks":10,"topic":"Arrays","question":"Given an array nums, rotate it to the right by k steps. Print the rotated array.","hint":"Reverse the whole array, then reverse first k elements, then reverse the rest.","test_cases":[{"input": "7\n1 2 3 4 5 6 7\n3", "expected_output": "5 6 7 1 2 3 4"}, {"input": "3\n-1 -100 3\n2", "expected_output": "3 -1 -100"}, {"input": "4\n1 2 3 4\n1", "expected_output": "4 1 2 3"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\nk = int(input())\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        int k = sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_thirdmax","title":"Third Maximum Number","marks":10,"topic":"Arrays","question":"Given an integer array nums, return the third distinct maximum. If it does not exist, return the maximum.","hint":"Track top 3 maximums using a sorted set or three variables.","test_cases":[{"input": "3\n3 2 1", "expected_output": "1"}, {"input": "3\n1 2", "expected_output": "2"}, {"input": "5\n2 2 3 1 4", "expected_output": "2"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_validpalindrome","title":"Valid Palindrome","marks":10,"topic":"Strings","question":"A phrase is a palindrome if, after converting all uppercase letters to lowercase and removing all non-alphanumeric characters, it reads the same forward and backward. Return true or false.","hint":"Filter alphanumeric, lowercase, compare with reverse.","test_cases":[{"input": "A man, a plan, a canal: Panama", "expected_output": "true"}, {"input": "race a car", "expected_output": "false"}, {"input": " ", "expected_output": "true"}],"template_python":"s = input()\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        String s = sc.nextLine();\n        // Write your solution here\n    }\n}"},
    {"id":"c_mergesorted","title":"Merge Sorted Array","marks":10,"topic":"Arrays","question":"Merge nums1 (with m elements + n zeros) and nums2 (n elements) into nums1 sorted. Print merged array.","hint":"Merge from the end using three pointers to avoid overwriting.","test_cases":[{"input": "3 3\n1 2 3 0 0 0\n2 5 6", "expected_output": "1 2 2 3 5 6"}, {"input": "1 0\n1\n", "expected_output": "1"}, {"input": "0 1\n0\n1", "expected_output": "1"}],"template_python":"m, n = map(int, input().split())\nnums1 = list(map(int, input().split()))\nif n > 0:\n    nums2 = list(map(int, input().split()))\nelse:\n    nums2 = []\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc=new Scanner(System.in);\n        int m=sc.nextInt(), n=sc.nextInt();\n        int[] nums1=new int[m+n]; for(int i=0;i<m+n;i++) nums1[i]=sc.nextInt();\n        int[] nums2=new int[n]; for(int i=0;i<n;i++) nums2[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_maxproduct","title":"Maximum Product Subarray","marks":10,"topic":"Arrays","question":"Given an integer array nums, find a subarray that has the largest product, and return the product.","hint":"Track both max and min product ending at each position (negative \u00d7 negative = positive).","test_cases":[{"input": "6\n2 3 -2 4", "expected_output": "6"}, {"input": "2\n-2 0 -1", "expected_output": "0"}, {"input": "3\n-2 3 -4", "expected_output": "24"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_minsubarray","title":"Minimum Size Subarray Sum","marks":10,"topic":"Arrays","question":"Given a target and array nums, return the minimum length of a subarray whose sum >= target. If none, return 0.","hint":"Sliding window: expand right, shrink left when sum >= target.","test_cases":[{"input": "7\n5\n2 3 1 2 4 3", "expected_output": "2"}, {"input": "4\n5\n1 4 4", "expected_output": "1"}, {"input": "11\n4\n1 1 1 1 1", "expected_output": "0"}],"template_python":"target = int(input())\nn = int(input())\nnums = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc=new Scanner(System.in);\n        int target=sc.nextInt(), n=sc.nextInt();\n        int[] nums=new int[n]; for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_productexcept","title":"Product of Array Except Self","marks":10,"topic":"Arrays","question":"Given integer array nums, return an array answer such that answer[i] equals the product of all elements except nums[i]. Must run in O(n) without using division.","hint":"Use left pass and right pass to compute products.","test_cases":[{"input": "4\n1 2 3 4", "expected_output": "24 12 8 6"}, {"input": "5\n-1 1 0 -3 3", "expected_output": "0 0 9 0 0"}, {"input": "3\n2 3 4", "expected_output": "12 8 6"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_container","title":"Container With Most Water","marks":10,"topic":"Arrays","question":"Given n non-negative integers as heights, find two lines that together with the x-axis form a container that holds the most water. Return the max water.","hint":"Two pointer approach: start from both ends, move the pointer with smaller height inward.","test_cases":[{"input": "9\n1 8 6 2 5 4 8 3 7", "expected_output": "49"}, {"input": "2\n1 1", "expected_output": "1"}, {"input": "6\n4 3 2 1 4", "expected_output": "16"}],"template_python":"n = int(input())\nheight = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] h = new int[n];\n        for(int i=0;i<n;i++) h[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_searchrotated","title":"Search in Rotated Sorted Array","marks":10,"topic":"Searching","question":"Given a rotated sorted array nums and a target value, return its index or -1 if not found. Must be O(log n).","hint":"Modified binary search: determine which half is sorted, then check if target is in that half.","test_cases":[{"input": "7\n4 5 6 7 0 1 2\n0", "expected_output": "4"}, {"input": "7\n4 5 6 7 0 1 2\n3", "expected_output": "-1"}, {"input": "1\n1\n0", "expected_output": "-1"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\ntarget = int(input())\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        int target = sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_combinationsum","title":"Combination Sum","marks":10,"topic":"Arrays","question":"Given array of distinct integers and a target, return all unique combinations that sum to target. Same number may be used multiple times. Print each combination on new line, sorted.","hint":"Backtracking: at each step, either include candidate or skip to next.","test_cases":[{"input": "4\n2 3 6 7\n7", "expected_output": "2 2 3\n7"}, {"input": "3\n2 3 5\n8", "expected_output": "2 2 2 2\n2 3 3\n3 5"}, {"input": "1\n2\n1", "expected_output": ""}],"template_python":"n = int(input())\ncandidates = list(map(int, input().split()))\ntarget = int(input())\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] c = new int[n];\n        for(int i=0;i<n;i++) c[i]=sc.nextInt();\n        int target = sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_nextperm","title":"Next Permutation","marks":10,"topic":"Arrays","question":"Rearrange nums into the next lexicographically greater permutation. If not possible, rearrange to smallest order. Print result.","hint":"Find rightmost element smaller than its next. Swap with next larger element to its right. Reverse suffix.","test_cases":[{"input": "3\n1 2 3", "expected_output": "1 3 2"}, {"input": "3\n3 2 1", "expected_output": "1 2 3"}, {"input": "3\n1 1 5", "expected_output": "1 5 1"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_firstlast","title":"Find First and Last Position in Sorted Array","marks":10,"topic":"Searching","question":"Given sorted array nums and target, return [first, last] positions. If not found, return [-1, -1].","hint":"Binary search twice: once biased to left, once biased to right.","test_cases":[{"input": "8\n5 7 7 8 8 10\n8", "expected_output": "3 4"}, {"input": "6\n5 7 7 8 8 10\n6", "expected_output": "-1 -1"}, {"input": "0\n\n0", "expected_output": "-1 -1"}],"template_python":"n = int(input())\nnums = list(map(int, input().split())) if n > 0 else []\ntarget = int(input())\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        int target = sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_3sum","title":"3Sum","marks":10,"topic":"Arrays","question":"Find all unique triplets in array that sum to 0. Print each triplet sorted on its own line. No duplicate triplets.","hint":"Sort array. Fix one element, use two pointers for remaining two. Skip duplicates.","test_cases":[{"input": "6\n-1 0 1 2 -1 -4", "expected_output": "-1 -1 2\n-1 0 1"}, {"input": "3\n0 1 1", "expected_output": ""}, {"input": "3\n0 0 0", "expected_output": "0 0 0"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_mergeintervals","title":"Merge Intervals","marks":10,"topic":"Arrays","question":"Given an array of intervals, merge all overlapping intervals. Print each merged interval.","hint":"Sort by start time. Merge if next start <= current end.","test_cases":[{"input": "4\n1 3\n2 6\n8 10\n15 18", "expected_output": "1 6\n8 10\n15 18"}, {"input": "2\n1 4\n4 5", "expected_output": "1 5"}, {"input": "3\n1 4\n2 3\n5 7", "expected_output": "1 4\n5 7"}],"template_python":"n = int(input())\nintervals = [list(map(int, input().split())) for _ in range(n)]\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[][] iv = new int[n][2];\n        for(int i=0;i<n;i++){iv[i][0]=sc.nextInt();iv[i][1]=sc.nextInt();}\n        // Write your solution here\n    }\n}"},
    {"id":"c_jumpgame","title":"Jump Game","marks":10,"topic":"Arrays","question":"Given array nums where nums[i] is max jump length from position i, return true if you can reach the last index.","hint":"Track maximum reachable index. If current index > max_reach, return false.","test_cases":[{"input": "5\n2 3 1 1 4", "expected_output": "true"}, {"input": "6\n3 2 1 0 4", "expected_output": "false"}, {"input": "3\n1 0 1", "expected_output": "false"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_firstbadver","title":"First Bad Version","marks":10,"topic":"Searching","question":"Given n versions and the first bad version k (1-indexed), find it using minimum API calls. You are given the bad version directly as input to simulate isBadVersion(v).","hint":"Binary search: if mid is bad, look left. If good, look right.","test_cases":[{"input": "5\n4", "expected_output": "4"}, {"input": "1\n1", "expected_output": "1"}, {"input": "10\n7", "expected_output": "7"}],"template_python":"n, bad = int(input()), int(input())\ndef isBadVersion(v): return v >= bad\n# Write your solution using isBadVersion(v)\n","template_java":"import java.util.*;\npublic class Main {\n    static int bad;\n    static boolean isBadVersion(int v){return v>=bad;}\n    public static void main(String[] args) {\n        Scanner sc=new Scanner(System.in);\n        int n=sc.nextInt(); bad=sc.nextInt();\n        // Write your solution here, print the first bad version\n    }\n}"},
    {"id":"c_searchinsert","title":"Search Insert Position","marks":10,"topic":"Searching","question":"Given sorted array and target, return index if found. If not, return index where it would be inserted to keep order.","hint":"Standard binary search. When not found, return left pointer.","test_cases":[{"input": "4\n1 3 5 6\n5", "expected_output": "2"}, {"input": "4\n1 3 5 6\n2", "expected_output": "1"}, {"input": "4\n1 3 5 6\n7", "expected_output": "4"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\ntarget = int(input())\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        int target = sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_findminrotated","title":"Find Minimum in Rotated Sorted Array","marks":10,"topic":"Searching","question":"Given a sorted rotated array of unique elements, find the minimum element. Must be O(log n).","hint":"Binary search: if mid > right, minimum is in right half. Else in left half.","test_cases":[{"input": "5\n3 4 5 1 2", "expected_output": "1"}, {"input": "6\n4 5 6 7 0 1 2", "expected_output": "0"}, {"input": "1\n11", "expected_output": "11"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_search2dmatrix","title":"Search a 2D Matrix","marks":10,"topic":"Searching","question":"Given m\u00d7n matrix (each row sorted, first element of each row > last of prev row) and target, return true/false.","hint":"Treat matrix as flattened sorted array. Binary search with mid/n and mid%n for row/col.","test_cases":[{"input": "3 4\n1 3 5 7\n10 11 16 20\n23 30 34 60\n3", "expected_output": "true"}, {"input": "3 4\n1 3 5 7\n10 11 16 20\n23 30 34 60\n13", "expected_output": "false"}, {"input": "1 1\n1\n1", "expected_output": "true"}],"template_python":"m, n = map(int, input().split())\nmatrix = [list(map(int, input().split())) for _ in range(m)]\ntarget = int(input())\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc=new Scanner(System.in);\n        int m=sc.nextInt(),n=sc.nextInt();\n        int[][]mat=new int[m][n];\n        for(int i=0;i<m;i++) for(int j=0;j<n;j++) mat[i][j]=sc.nextInt();\n        int target=sc.nextInt();\n        // Write your solution here\n    }\n}"},
    {"id":"c_firstmisspos","title":"First Missing Positive","marks":10,"topic":"Arrays","question":"Given unsorted integer array nums, return the smallest missing positive integer. Must run in O(n) time and O(1) extra space.","hint":"Place each number in its correct position (num at index num-1). Then scan for first index where nums[i] != i+1.","test_cases":[{"input": "3\n1 2 0", "expected_output": "3"}, {"input": "4\n3 4 -1 1", "expected_output": "2"}, {"input": "3\n7 8 9", "expected_output": "1"}],"template_python":"n = int(input())\nnums = list(map(int, input().split()))\n# Write your solution here\n","template_java":"import java.util.*;\npublic class Main {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int n = sc.nextInt();\n        int[] nums = new int[n];\n        for(int i=0;i<n;i++) nums[i]=sc.nextInt();\n        // Write your solution here\n    }\n}"}
]

import random

def get_random_questions_for_student(seed=None):
    """Pick random questions from each bank — different per student.
    MCQs are sampled evenly across sections/topics for balanced paper."""
    rng = random.Random(seed)

    # Group MCQs by section for balanced sampling
    from collections import defaultdict
    by_section = defaultdict(list)
    for q in MCQ_BANK:
        by_section[q.get("section","General")].append(q)

    mcqs = []
    sections = list(by_section.keys())
    per_section = max(1, MCQ_PER_EXAM // len(sections))
    remaining   = MCQ_PER_EXAM

    for sec in sections:
        pool   = by_section[sec]
        take   = min(per_section, len(pool), remaining)
        mcqs  += rng.sample(pool, take)
        remaining -= take
        if remaining <= 0:
            break

    # Fill up to MCQ_PER_EXAM if needed
    if remaining > 0:
        used_ids = {q["id"] for q in mcqs}
        leftover = [q for q in MCQ_BANK if q["id"] not in used_ids]
        if leftover:
            mcqs += rng.sample(leftover, min(remaining, len(leftover)))

    rng.shuffle(mcqs)

    sqls    = rng.sample(SQL_BANK,    min(SQL_PER_EXAM,    len(SQL_BANK)))
    codings = rng.sample(CODING_BANK, min(CODE_PER_EXAM,   len(CODING_BANK)))
    return {"mcq": mcqs, "sql": sqls, "coding": codings}

# Legacy: keep QUESTIONS pointing to MCQ bank for backwards compat
QUESTIONS = MCQ_BANK[:10]


# ==============================================================================
# MODEL LOADING
# ==============================================================================
# MODEL LOADING
# ==============================================================================
def load_models():
    global _cnn_model, _rf_model, _scaler, _yolo_model
    global _model_type, _yolo_fine, YOLO_PERSON_ID, YOLO_PHONE_ID

    # Scaler
    for nm in ["scaler_v6.pkl","scaler.pkl"]:
        p = os.path.join(MODEL_DIR, nm)
        if os.path.exists(p):
            try:
                with open(p,"rb") as f: _scaler = pickle.load(f)
                print(f"[MODEL] ✅ Scaler loaded: {p}"); break
            except Exception as e: print(f"[MODEL] Scaler err: {e}")

    # CNN
    if HAS_TF:
        for nm in ["cnn_phone.h5","proctor_cnn_v6.keras","best_proctor_v6.keras",
                   "proctor_cnn.keras","best_proctor.keras"]:
            p = os.path.join(MODEL_DIR, nm)
            if os.path.exists(p):
                try:
                    _cnn_model = tf.keras.models.load_model(p)
                    _model_type = "keras"
                    print(f"[MODEL] ✅ CNN: {p}  ({_cnn_model.count_params():,} params)")
                    break
                except Exception as e: print(f"[MODEL] CNN {nm} err: {e}")

    # RandomForest
    for nm in ["proctor_rf_v6.pkl","proctor_rf.pkl"]:
        p = os.path.join(MODEL_DIR, nm)
        if os.path.exists(p):
            try:
                with open(p,"rb") as f: _rf_model = pickle.load(f)
                if _model_type == "heuristic": _model_type = "sklearn"
                print(f"[MODEL] ✅ RF: {p}")
                break
            except Exception as e: print(f"[MODEL] RF err: {e}")

    # YOLOv8 — prefer fine-tuned, fall back to COCO
    if HAS_YOLO:
        fine_path = os.path.join(MODEL_DIR,"yolo_proctor.pt")
        if os.path.exists(fine_path):
            try:
                _yolo_model = UltralyticsYOLO(fine_path)
                _yolo_model(np.zeros((640,640,3),np.uint8), verbose=False)
                _yolo_fine = True
                YOLO_PERSON_ID = YOLO_PERSON_FINE
                YOLO_PHONE_ID  = YOLO_PHONE_FINE
                print(f"[YOLO] ✅ Fine-tuned model: {fine_path}  (phone=class{YOLO_PHONE_ID})")
            except Exception as e:
                print(f"[YOLO] Fine-tuned failed: {e}")
                _yolo_model = None

        if _yolo_model is None:
            try:
                _yolo_model = UltralyticsYOLO("yolov8n.pt")
                _yolo_model(np.zeros((640,640,3),np.uint8), verbose=False)
                _yolo_fine = False
                YOLO_PERSON_ID = YOLO_PERSON_COCO
                YOLO_PHONE_ID  = YOLO_PHONE_COCO
                print(f"[YOLO] ✅ COCO yolov8n.pt  (phone=class67, person=class0)")
            except Exception as e:
                print(f"[YOLO] Failed: {e}"); _yolo_model = None

    if _model_type == "heuristic":
        print("[MODEL] ⚠  Running heuristic mode — run training.py first")



def _boxes_non_overlapping(boxes, iou_threshold=0.25):
    """Return True only if no two boxes overlap more than iou_threshold.
    Prevents counting the same person twice (e.g. partial body + full body)."""
    if len(boxes) < 2:
        return False  # need at least 2
    for i in range(len(boxes)):
        for j in range(i+1, len(boxes)):
            b1, b2 = boxes[i], boxes[j]
            # Intersection
            ix1 = max(b1[0], b2[0]); iy1 = max(b1[1], b2[1])
            ix2 = min(b1[2], b2[2]); iy2 = min(b1[3], b2[3])
            if ix2 <= ix1 or iy2 <= iy1:
                inter = 0.0
            else:
                inter = (ix2-ix1)*(iy2-iy1)
            area1 = max((b1[2]-b1[0])*(b1[3]-b1[1]), 1e-6)
            area2 = max((b2[2]-b2[0])*(b2[3]-b2[1]), 1e-6)
            iou   = inter / (area1 + area2 - inter + 1e-6)
            if iou > iou_threshold:
                return False  # overlapping — likely same person detected twice
    return True


# ==============================================================================
# YOLO INFERENCE
# ==============================================================================
def run_yolo(img_bgr):
    """Returns: phone_conf, phone_boxes, person_count, person_boxes, person_confs
    All boxes normalised [x1,y1,x2,y2,conf] in [0,1]."""
    if _yolo_model is None:
        return 0.0, [], 0, [], []
    try:
        H, W = img_bgr.shape[:2]
        # Use low conf so small phones aren't discarded by NMS
        res = _yolo_model(img_bgr, verbose=False, conf=0.10)[0]
        phone_conf   = 0.0
        phone_boxes  = []
        person_count = 0
        person_boxes = []
        person_confs = []
        for box in res.boxes:
            cls  = int(box.cls[0])
            conf = float(box.conf[0])
            x1,y1,x2,y2 = box.xyxy[0].tolist()
            nb = [x1/W, y1/H, x2/W, y2/H, conf]
            if cls == YOLO_PHONE_ID:
                # COCO class 67 gives lower confidence for phones — accept >= 0.15
                if conf >= 0.15:
                    phone_conf = max(phone_conf, conf)
                    phone_boxes.append(nb)
            elif cls == YOLO_PERSON_ID:
                # Very strict filters — only count real, large, upright persons in frame
                if conf < 0.65:                 # raised from 0.55 — cuts background detections
                    continue
                box_h = (y2 - y1) / H
                box_w = (x2 - x1) / W
                if box_h * box_w < 0.12:        # raised from 0.08 — must cover 12% of frame
                    continue
                if (box_h / max(box_w, 0.01)) < 0.8:   # must be clearly taller than wide (upright person)
                    continue
                # Extra: bounding box must reach at least 30% of frame height to be a real person
                if box_h < 0.30:
                    continue
                person_count += 1
                person_boxes.append(nb)
                person_confs.append(conf)
        return phone_conf, phone_boxes, person_count, person_boxes, person_confs
    except Exception as e:
        print(f"[YOLO] {e}")
        return 0.0, [], 0, [], []


# ==============================================================================
# CNN + RF ENSEMBLE
# ==============================================================================
def predict_ensemble(features):
    try:
        f = features.reshape(1,-1).astype(np.float32)
        if _scaler is not None:
            f = _scaler.transform(f)

        cnn_p = rf_p = None
        if _cnn_model is not None and _model_type == "keras":
            cnn_p = _cnn_model.predict(f.reshape(1,NUM_FEATURES,1), verbose=0)[0]
        if _rf_model is not None:
            rp = _rf_model.predict_proba(f)[0]
            if len(rp) < NUM_CLASSES:
                rp = np.pad(rp,(0,NUM_CLASSES-len(rp)))
            rf_p = rp

        if cnn_p is not None and rf_p is not None:
            probs = 0.60*cnn_p + 0.40*rf_p
        elif cnn_p is not None: probs = cnn_p
        elif rf_p  is not None: probs = rf_p
        else:
            # Heuristic fallback
            r = f[0]
            ph = float(np.clip(r[0]  if len(r)>0  else 0,0,1))
            mu = float(np.clip(r[13] if len(r)>13 else 0,0,1))
            vp = float(np.clip(r[27] if len(r)>27 else 0,0,1))
            id_ = float(np.clip(1.0-r[43] if len(r)>43 else 0,0,1))
            base = [max(0.0,1.0-ph*1.2-mu*1.1-vp*1.1-id_*1.3),
                    min(1.0,ph*1.2), min(1.0,mu*1.1),
                    min(1.0,vp*1.1), min(1.0,id_*1.3)]
            s = sum(base)+1e-9
            probs = np.array([b/s for b in base],dtype=np.float32)

        probs = np.clip(np.array(probs,dtype=np.float32),0,1)
        pred  = int(np.argmax(probs))
        thr   = [0.50, 0.40, 0.40, 0.38, 0.40]
        if probs[pred] < thr[pred]: pred = 0

        return {"pred_class":pred, "class_name":CLASS_NAMES[pred],
                "confidence":float(probs[pred]),
                "probabilities":{CLASS_NAMES[i]:round(float(p),4) for i,p in enumerate(probs)}}
    except Exception as e:
        print(f"[PREDICT] {e}")
        return {"pred_class":0,"class_name":"Normal","confidence":0.9,
                "probabilities":{n:0.2 for n in CLASS_NAMES}}


# ==============================================================================
# FEATURE EXTRACTION  (72 features)
# ==============================================================================
def _fill_identity(feat, audio_rms, audio_zcr, ref_embed,
                   face_mean=0.5, face_std=0.15):
    if ref_embed is None:
        feat[36]=0.05; feat[37]=0.92; feat[40]=0.05; feat[41]=0.90; feat[43]=0.94
        return
    rfm = ref_embed.get("face_mean",0.5)
    rfs = ref_embed.get("face_std", 0.15)
    fd  = float(np.clip(abs(face_mean-rfm)*2.0 + abs(face_std-rfs)*1.0, 0,1))
    fm  = float(np.clip(1.0-fd*1.5, 0,1))
    rrms = ref_embed.get("audio_rms",0.05)
    rzcr = ref_embed.get("audio_zcr",0.30)
    if audio_rms < 0.02 and rrms < 0.02:
        vd = 0.05; vm = 0.90
    else:
        vd = float(np.clip(abs(audio_rms-rrms)*3.0 + abs(audio_zcr-rzcr)*1.0, 0,1))
        vm = float(np.clip(1.0-vd*1.3, 0,1))
    ic = fm*0.6 + vm*0.4
    feat[36]=fd;  feat[37]=fm;  feat[38]=max(0.0,fm-0.1); feat[39]=max(0.0,1.0-fd)
    feat[40]=vd;  feat[41]=vm;  feat[42]=max(0.0,vm-0.1)
    feat[43]=ic;  feat[44]=min(1.0,fm+0.05)
    feat[45]=float(abs(face_mean-rfm)*2); feat[46]=float(abs(face_std-rfs)*3)
    feat[47]=float(ic < 0.40)


def extract_features(img_bytes, audio_rms=0.0, audio_zcr=0.0,
                     audio_peak=0.0, ref_embed=None):
    feat = np.zeros(NUM_FEATURES, dtype=np.float32)
    feat[4]=0.85; feat[37]=0.92; feat[41]=0.90; feat[43]=0.94
    feat[60]=0.90; feat[61]=0.88

    phone_boxes_out  = []
    person_boxes_out = []
    person_count_out = 0

    if not HAS_CV:
        raw = np.frombuffer(img_bytes,np.uint8)
        bri = float(raw[-min(4000,len(raw)):].mean())/255.0
        feat[1]=bri; feat[0]=max(0.0,(bri-0.35)*3.0); feat[12]=feat[0]
        feat[24]=float(np.clip(audio_rms*4,0,1))
        feat[25]=float(np.clip(audio_zcr*2,0,1))
        feat[27]=float(np.clip(audio_rms*3,0,1))
        _fill_identity(feat, audio_rms, audio_zcr, ref_embed)
        return feat, [], [], 0

    try:
        nparr   = np.frombuffer(img_bytes, np.uint8)
        img_bgr = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if img_bgr is None:
            return feat, [], [], 0
        H, W = img_bgr.shape[:2]
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        hsv  = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)
        gf   = gray.astype(np.float32)/255.0

        # ── YOLOv8 inference ──────────────────────────────────────────────────
        phone_yolo_conf, phone_boxes, person_count, person_boxes, person_confs = \
            run_yolo(img_bgr)
        phone_boxes_out  = phone_boxes
        person_boxes_out = person_boxes
        person_count_out = person_count

        # ── [0-11] Phone features ──────────────────────────────────────────────
        # Scan 3 regions — phone can appear anywhere (chest, desk, face-level)
        best_hp = 0.0
        for rgn in [gf[int(H*0.45):,:], gf[int(H*0.20):int(H*0.80),:], gf]:
            _rm  = float(rgn.mean()); _rs = float(rgn.std())
            _pb  = float(np.clip((_rm-0.28)/0.55, 0, 1))
            _pu  = float(np.clip(1.0-_rs/0.20,    0, 1))
            _hp  = float(np.clip((_pb*0.65+_pu*0.35-0.25)*2, 0, 1))
            best_hp = max(best_hp, _hp)

        lo  = gf[int(H*0.45):,:]
        lbm = float(lo.mean()); lbs = float(lo.std())
        pb  = float(np.clip((lbm-0.28)/0.55, 0, 1))
        pu  = float(np.clip(1.0-lbs/0.20,    0, 1))
        lo8 = (lo*255).astype(np.uint8)
        edg = cv2.Canny(lo8, 50, 150)
        rs  = float(np.clip(edg.mean()/128, 0, 1))
        hp  = best_hp
        # Weight YOLO heavily when it has a confident phone detection
        if phone_yolo_conf >= 0.15:
            pc = float(np.clip(phone_yolo_conf*0.90 + hp*0.10, 0, 1))
        elif phone_yolo_conf > 0.0:
            pc = float(np.clip(phone_yolo_conf*0.75 + hp*0.25, 0, 1))
        else:
            pc = hp

        feat[0]=phone_yolo_conf; feat[1]=lbm; feat[2]=lbs; feat[3]=pb; feat[4]=pu
        feat[5]=rs; feat[6]=float(np.clip((lbm-0.52)*2,0,1))
        feat[7]=pc; feat[8]=lbm if phone_yolo_conf>0.35 else pb*pu
        feat[9]=pc*0.9; feat[10]=pc*0.85; feat[11]=rs; feat[12]=pc

        # ── [12-23] Multi-person features ──────────────────────────────────────
        feat[16] = float(np.clip(person_count,0,5))
        feat[17] = float(np.clip(person_count,0,5))
        feat[20] = float(np.clip(person_count/4.0,0,1))
        feat[21] = float(np.clip(max(0,person_count-1)*0.3,0,1))

        if person_count >= 2:
            f2c = float(person_confs[1]) if len(person_confs)>1 else 0.85
            if len(person_boxes) > 1:
                b = person_boxes[1]
                feat[14] = float(np.clip((b[2]-b[0])*(b[3]-b[1]),0,0.25))
                feat[15] = f2c
            feat[13] = f2c; feat[23] = f2c
        else:
            # Haar fallback — extremely strict; only fire when YOLO missed a very clear 2nd face
            casc = cv2.CascadeClassifier(
                cv2.data.haarcascades+"haarcascade_frontalface_default.xml")
            # minNeighbors=15 (was 10) — eliminates almost all false detections
            # minSize=(110,110) — only large, close faces qualify
            fces = casc.detectMultiScale(gray, 1.1, 15, minSize=(110,110))
            if len(fces) >= 2:
                _,_,fw2,fh2 = fces[1]
                r2 = (fw2*fh2)/(W*H)
                # 2nd face must cover 8% of frame (was 5%) — rules out far-background faces
                if r2 >= 0.08:
                    feat[13] = 0.78
                    feat[14] = float(np.clip(r2,0,0.25))

        # Skin tone detection
        sk1 = cv2.inRange(hsv,np.array([0,50,80],np.uint8),np.array([20,220,255],np.uint8))
        sk2 = cv2.inRange(hsv,np.array([0,30,60],np.uint8),np.array([18,180,200],np.uint8))
        sk  = cv2.morphologyEx(cv2.bitwise_or(sk1,sk2),
                cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_ELLIPSE,(7,7)))
        feat[19] = float(np.clip(float(sk.sum())/(H*W*255)*3.5,0,1))
        feat[22] = 0.0   # disabled — caused false multi-person from background colours

        # ── Face ROI → head pose + identity ────────────────────────────────────
        face_mean_val = 0.5; face_std_val = 0.15
        casc2 = cv2.CascadeClassifier(
            cv2.data.haarcascades+"haarcascade_frontalface_default.xml")
        fcs = casc2.detectMultiScale(gray,1.1,4,minSize=(40,40))
        if len(fcs) > 0:
            x,y,fw,fh = fcs[0]
            cx = (x+fw/2)/W; cy = (y+fh/2)/H
            roi = gray[y:y+fh,x:x+fw].astype(np.float32)/255.0
            face_mean_val = float(roi.mean()); face_std_val = float(roi.std())
            yaw   = float(np.clip((cx-0.5)*2.0,-1,1))
            pitch = float(np.clip((cy-0.35)*1.5,-0.5,0.5))
            feat[48]=yaw; feat[49]=pitch; feat[50]=0.0
            feat[51]=abs(yaw); feat[52]=abs(pitch)
            feat[53]=float(np.clip(fw/W,0,1)); feat[54]=float(np.clip(fh/H,0,1))
            feat[55]=abs(yaw)*1.8   # side_view_score
            feat[56]=abs(yaw)*0.6+abs(pitch)*0.4
            mid = x+fw//2
            lm=float(gf[:,:mid].mean()); rm=float(gf[:,mid:].mean())
            feat[57]=float(np.clip(1.0-abs(lm-rm)*3,0,1))
            feat[58]=float(np.clip(1.0-abs(yaw)*1.6,0,1))
            feat[60]=float(np.clip(1.0-feat[56]*2-feat[51],0,1))
            feat[61]=float(np.clip(feat[60]*0.7+feat[58]*0.3,0,1))
            lip = gray[y+int(fh*0.65):y+fh, x:x+fw]
            if lip.size > 0:
                feat[30]=float(np.clip(np.abs(cv2.Laplacian(lip,cv2.CV_32F)).mean()/30,0,1))
        else:
            feat[60] = 0.25  # no face visible

        # ── [24-35] Audio features ─────────────────────────────────────────────
        feat[24]=float(np.clip(audio_rms*4,0,1))
        feat[25]=float(np.clip(audio_zcr*2,0,1))
        feat[26]=float(np.clip(audio_peak*3,0,1))
        vp = float(np.clip(audio_rms*2.5+audio_zcr*0.5,0,1))
        feat[27]=vp; feat[28]=float(np.clip(audio_rms*1.5,0,1))
        feat[29]=float(np.clip(1.0-audio_rms,0,1))
        feat[31]=float(np.clip(vp*0.9,0,1)); feat[32]=float(np.clip(vp*0.85,0,1))
        feat[34]=float(np.clip(audio_rms*1.2,0,1)); feat[35]=float(np.clip(audio_rms*1.8,0,1))

        # ── [36-47] Identity features ──────────────────────────────────────────
        _fill_identity(feat, audio_rms, audio_zcr, ref_embed, face_mean_val, face_std_val)

        # ── [60-71] Motion / composite ─────────────────────────────────────────
        lap = cv2.Laplacian(gray,cv2.CV_32F)
        feat[62]=float(np.clip(np.abs(lap).mean()/30,0,1))
        feat[63]=float(np.clip(lap.std()/50,0,1))
        feat[64]=float(np.clip(np.abs(cv2.Laplacian(gray[:H//2,:],cv2.CV_32F)).mean()/30,0,1))
        feat[65]=float(np.clip(np.abs(cv2.Laplacian(gray[H//2:,:],cv2.CV_32F)).mean()/30,0,1))
        feat[70]=float(np.clip(pc*0.35+feat[13]*0.30+vp*0.15+(1-feat[43])*0.20,0,1))
        feat[71]=float(np.clip(pc*0.40+feat[13]*0.35+feat[51]*0.15+(feat[20]-0.25)*0.10,0,1))

    except Exception:
        print(f"[FEAT] {traceback.format_exc()}")

    return feat, phone_boxes_out, person_boxes_out, person_count_out


# ==============================================================================
# HELPERS
# ==============================================================================
def cheating_score(s):
    sc  = min(s.get("phone_count",        0)*12, 35)
    sc += min(s.get("multi_person_count", 0)*14, 30)
    sc += min(s.get("voice_count",        0)*8,  20)
    sc += min(s.get("face_mismatch_count",0)*18, 35)
    sc += min(s.get("tab_switch_count",   0)*15, 30)
    return min(int(sc), 100)


def session_expired(s):
    """Return True if an active session has elapsed beyond EXAM_DURATION + 5 min grace."""
    if s.get("status") != "active":
        return False
    try:
        started = datetime.fromisoformat(s["start_time"])
        elapsed = (datetime.now() - started).total_seconds()
        return elapsed > (EXAM_DURATION + 300)
    except Exception:
        return False


def new_session(username, email):
    sid = str(uuid.uuid4())
    SESSIONS[sid] = {
        "session_id":sid,"username":username,"email":email,
        "start_time":datetime.now().isoformat(),"end_time":None,
        "answers":{},"sql_answers":{},"code_answers":{},
        "score":None,"mcq_score":0,"sql_score":0,"code_score":0,
        "mcq_total":0,"sql_total":0,"code_total":0,"total_marks":0,
        "cheating_score":0,"warnings":[],
        "phone_count":0,"multi_person_count":0,"voice_count":0,
        "identity_count":0,"face_mismatch_count":0,"side_view_count":0,"tab_switch_count":0,
        "screenshots":[],"transcript":[],"ref_embed":None,"enrolled":False,
        "status":"active",
        "terminated":False,"terminate_reason":"","terminate_cause":"","terminate_time":None,
        "questions":None,"invite_token":None,
    }
    return sid


def admin_required(f):
    @wraps(f)
    def dec(*a,**kw):
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin_login"))
        return f(*a,**kw)
    return dec


def student_required(f):
    @wraps(f)
    def dec(*a,**kw):
        if not session.get("exam_sid"):
            return redirect(url_for("student_login"))
        return f(*a,**kw)
    return dec



# ==============================================================================
# EMAIL & EXAM LINK HELPERS
# ==============================================================================
EXAM_INVITES: dict = {}   # token -> {email, name, expires, questions_seed, used}

def _smtp_send(to_addr, subject, html_body, from_label):
    """Internal: send via SMTP (works for Gmail, Outlook, Brevo, SendGrid SMTP, etc.)"""
    cfg  = EMAIL_CONFIG
    user = cfg.get("smtp_user","").strip()
    pswd = cfg.get("smtp_pass","").strip()
    host = cfg.get("smtp_host","").strip()
    port = int(cfg.get("smtp_port", 587))
    from_email = cfg.get("from_email","").strip() or user
    if not user or not pswd or not host:
        return False, "SMTP credentials incomplete — fill in username and password."
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = f"{from_label} <{from_email}>"
        msg["To"]      = to_addr
        msg.attach(MIMEText(html_body, "html"))
        if cfg.get("use_ssl"):
            with smtplib.SMTP_SSL(host, port, timeout=15) as srv:
                srv.login(user, pswd)
                srv.sendmail(from_email, to_addr, msg.as_string())
        else:
            with smtplib.SMTP(host, port, timeout=15) as srv:
                srv.ehlo(); srv.starttls(); srv.ehlo()
                srv.login(user, pswd)
                srv.sendmail(from_email, to_addr, msg.as_string())
        return True, ""
    except smtplib.SMTPAuthenticationError as e:
        prov = EMAIL_CONFIG.get("provider","smtp")
        hints = {
            "brevo":    "Brevo: use the SMTP Password from brevo.com → SMTP & API → SMTP (not your account password).",
            "sendgrid": "SendGrid: username must be 'apikey' and password is the API key.",
            "gmail":    "Gmail: use an App Password, NOT your login password. Google Account → Security → App Passwords.",
        }
        return False, hints.get(prov, f"Auth failed: {e}")
    except Exception as e:
        return False, str(e)


def _brevo_api_send(to_addr, to_name, subject, html_body, from_label, from_email):
    """Send via Brevo HTTP API (most reliable, no SMTP needed)."""
    import urllib.request, json as _json
    api_key = EMAIL_CONFIG.get("api_key","").strip()
    if not api_key:
        return False, "Brevo API key not set."
    payload = _json.dumps({
        "sender":  {"name": from_label, "email": from_email},
        "to":      [{"email": to_addr, "name": to_name}],
        "subject": subject,
        "htmlContent": html_body,
    }).encode()
    req = urllib.request.Request(
        "https://api.brevo.com/v3/smtp/email",
        data=payload,
        headers={"api-key": api_key, "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return True, ""
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        return False, f"Brevo API error {e.code}: {body[:200]}"
    except Exception as e:
        return False, str(e)


def _sendgrid_api_send(to_addr, to_name, subject, html_body, from_label, from_email):
    """Send via SendGrid HTTP API."""
    import urllib.request, json as _json
    api_key = EMAIL_CONFIG.get("api_key","").strip()
    if not api_key:
        return False, "SendGrid API key not set."
    payload = _json.dumps({
        "personalizations": [{"to": [{"email": to_addr, "name": to_name}]}],
        "from": {"email": from_email, "name": from_label},
        "subject": subject,
        "content": [{"type": "text/html", "value": html_body}],
    }).encode()
    req = urllib.request.Request(
        "https://api.sendgrid.com/v3/mail/send",
        data=payload,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return True, ""
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        return False, f"SendGrid error {e.code}: {body[:200]}"
    except Exception as e:
        return False, str(e)


def send_email(to_addr, subject, html_body, to_name=""):
    """Send HTML email using whichever provider is configured."""
    cfg = EMAIL_CONFIG
    if not cfg.get("enabled"):
        return False, "Email not configured. Go to Admin → ⚙️ Settings → Email Setup."

    provider   = cfg.get("provider", "smtp").lower()
    from_label = cfg.get("from_name","ProctorAI Exams").strip() or "ProctorAI Exams"
    from_email = (cfg.get("from_email","").strip()
                  or cfg.get("smtp_user","").strip()
                  or "noreply@example.com")

    print(f"[EMAIL] Sending via {provider} to {to_addr}")

    # ── Brevo API (recommended) ──────────────────────────────────────────────
    if provider == "brevo" and cfg.get("api_key","").strip():
        ok, err = _brevo_api_send(to_addr, to_name or to_addr, subject, html_body, from_label, from_email)
        if not ok:
            print(f"[EMAIL] Brevo API failed ({err}), trying SMTP fallback…")
            ok, err = _smtp_send(to_addr, subject, html_body, from_label)
        return ok, err

    # ── SendGrid API ─────────────────────────────────────────────────────────
    if provider == "sendgrid" and cfg.get("api_key","").strip():
        return _sendgrid_api_send(to_addr, to_name or to_addr, subject, html_body, from_label, from_email)

    # ── SMTP (Gmail, Outlook, Brevo SMTP, Mailgun SMTP, custom) ─────────────
    ok, err = _smtp_send(to_addr, subject, html_body, from_label)
    if not ok: print(f"[EMAIL] SMTP error: {err}")
    return ok, err


def test_email_connection():
    """Test connection for whichever provider is configured."""
    cfg      = EMAIL_CONFIG
    provider = cfg.get("provider","smtp").lower()

    # Brevo API test — list senders endpoint (lightweight)
    if provider == "brevo" and cfg.get("api_key","").strip():
        import urllib.request
        req = urllib.request.Request(
            "https://api.brevo.com/v3/senders",
            headers={"api-key": cfg["api_key"].strip()},
            method="GET",
        )
        try:
            with urllib.request.urlopen(req, timeout=10):
                return True, "Brevo API key valid — ready to send emails! ✅"
        except urllib.error.HTTPError as e:
            return False, f"Brevo API error {e.code} — check your API key."
        except Exception as e:
            return False, str(e)

    # SendGrid test — get user profile
    if provider == "sendgrid" and cfg.get("api_key","").strip():
        import urllib.request
        req = urllib.request.Request(
            "https://api.sendgrid.com/v3/user/profile",
            headers={"Authorization": f"Bearer {cfg['api_key'].strip()}"},
            method="GET",
        )
        try:
            with urllib.request.urlopen(req, timeout=10):
                return True, "SendGrid API key valid — ready to send! ✅"
        except urllib.error.HTTPError as e:
            return False, f"SendGrid error {e.code} — check your API key."
        except Exception as e:
            return False, str(e)

    # SMTP test
    host = cfg.get("smtp_host","").strip()
    port = int(cfg.get("smtp_port", 587))
    user = cfg.get("smtp_user","").strip()
    pswd = cfg.get("smtp_pass","").strip()
    if not host or not user or not pswd:
        return False, "Credentials not set — fill in all fields."
    try:
        if cfg.get("use_ssl"):
            with smtplib.SMTP_SSL(host, port, timeout=10) as srv:
                srv.login(user, pswd)
        else:
            with smtplib.SMTP(host, port, timeout=10) as srv:
                srv.ehlo(); srv.starttls(); srv.ehlo(); srv.login(user, pswd)
        return True, f"Connected to {host}:{port} ✅"
    except smtplib.SMTPAuthenticationError:
        prov = cfg.get("provider","smtp")
        hints = {
            "brevo":    "Use the SMTP Password from brevo.com → SMTP & API (not your account password).",
            "sendgrid": "Username must be 'apikey' and password is your SendGrid API key.",
            "gmail":    "Use a Gmail App Password, not your login password.",
        }
        return False, hints.get(prov, "Auth failed — check username and password.")
    except Exception as e:
        return False, str(e)

def send_exam_invite_email(name, email, token):
    """Send exam invitation with unique link."""
    link = f"{EXAM_BASE_URL}/exam/start/{token}"
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:auto;background:#f8faff;padding:30px;border-radius:12px">
      <div style="background:linear-gradient(135deg,#1a1a2e,#16213e);padding:24px;border-radius:10px;text-align:center;margin-bottom:20px">
        <h1 style="color:white;margin:0;font-size:22px">🎓 ProctorAI Exam Invitation</h1>
      </div>
      <p style="font-size:15px;color:#333">Dear <b>{name}</b>,</p>
      <p style="font-size:14px;color:#555">You have been invited to take the online assessment. Your exam contains:</p>
      <ul style="font-size:14px;color:#555">
        <li>📝 <b>Part A:</b> {MCQ_PER_EXAM} Multiple Choice Questions</li>
        <li>🗄️ <b>Part B:</b> {SQL_PER_EXAM} SQL Questions</li>
        <li>💻 <b>Part C:</b> {CODE_PER_EXAM} Coding Problems (with live compiler)</li>
      </ul>
      <div style="text-align:center;margin:28px 0">
        <a href="{link}" style="background:#2563eb;color:white;padding:14px 32px;border-radius:8px;text-decoration:none;font-size:16px;font-weight:bold">Start My Exam →</a>
      </div>
      <p style="font-size:12px;color:#888">This link expires in {EXAM_LINK_EXPIRY} hours. Use a laptop/desktop with webcam.</p>
      <p style="font-size:12px;color:#aaa;text-align:center">ProctorAI Proctoring System</p>
    </div>"""
    return send_email(email, "📝 Your Exam Invitation — ProctorAI", html)

def send_result_email(name, email, session_data, questions_data):
    """Send exam result summary to student."""
    score     = session_data.get("score", 0)
    total     = session_data.get("total_marks", 0)
    pct       = round(score/total*100, 1) if total else 0
    cs        = session_data.get("cheating_score", 0)
    terminated= session_data.get("terminated", False)
    status_txt= "⛔ TERMINATED" if terminated else ("✅ PASS" if pct >= 50 else "❌ FAIL")
    color     = "#d62828" if terminated else ("#16a34a" if pct >= 50 else "#ea580c")

    mcq_score  = session_data.get("mcq_score", 0)
    sql_score  = session_data.get("sql_score", 0)
    code_score = session_data.get("code_score", 0)
    mcq_total  = session_data.get("mcq_total", 0)
    sql_total  = session_data.get("sql_total", 0)
    code_total = session_data.get("code_total", 0)

    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:620px;margin:auto;background:#f8faff;padding:30px;border-radius:12px">
      <div style="background:linear-gradient(135deg,#1a1a2e,#16213e);padding:24px;border-radius:10px;text-align:center;margin-bottom:20px">
        <h1 style="color:white;margin:0;font-size:22px">📊 Your Exam Result</h1>
      </div>
      <p style="font-size:15px;color:#333">Dear <b>{name}</b>, here is your performance summary:</p>
      <div style="background:white;border-radius:10px;padding:20px;margin:16px 0;border:2px solid {color}">
        <div style="text-align:center">
          <div style="font-size:36px;font-weight:bold;color:{color}">{score}/{total}</div>
          <div style="font-size:18px;color:{color};font-weight:bold">{status_txt}</div>
          <div style="font-size:14px;color:#888">Overall Score: {pct}%</div>
        </div>
      </div>
      <table style="width:100%;border-collapse:collapse;font-size:14px;margin:12px 0">
        <tr style="background:#e8f0fe"><th style="padding:8px;text-align:left">Section</th><th>Score</th><th>Total</th></tr>
        <tr><td style="padding:8px">📝 Part A — MCQ</td><td style="text-align:center">{mcq_score}</td><td style="text-align:center">{mcq_total}</td></tr>
        <tr style="background:#f8f9ff"><td style="padding:8px">🗄️ Part B — SQL</td><td style="text-align:center">{sql_score}</td><td style="text-align:center">{sql_total}</td></tr>
        <tr><td style="padding:8px">💻 Part C — Coding</td><td style="text-align:center">{code_score}</td><td style="text-align:center">{code_total}</td></tr>
      </table>
      <div style="font-size:13px;color:#666;padding:12px;background:#fff3cd;border-radius:8px;margin-top:12px">
        ⚠ Integrity Score: <b>{cs}%</b> {"— Violations detected" if cs > 20 else "— Good"}
      </div>
      <p style="font-size:12px;color:#aaa;text-align:center;margin-top:20px">ProctorAI Proctoring System</p>
    </div>"""
    return send_email(email, "📊 Your Exam Result — ProctorAI", html)


# ==============================================================================
# SQL EVALUATOR
# ==============================================================================
def run_sql_query(setup_sql, user_query, expected_output, expected_cols):
    """Run user SQL against in-memory SQLite, compare with expected output."""
    result = {"passed": False, "output": [], "expected": expected_output,
              "error": "", "cols": []}
    try:
        con = sqlite3.connect(":memory:")
        cur = con.cursor()
        cur.executescript(setup_sql)
        cur.execute(user_query)
        rows = cur.fetchall()
        cols = [d[0].lower() for d in cur.description] if cur.description else []
        output = [list(row) for row in rows]
        result["output"] = output
        result["cols"]   = cols
        # Compare: sort both for order-independent check
        def norm(lst): return sorted([sorted(str(v) for v in r) for r in lst])
        result["passed"] = norm(output) == norm(expected_output)
        con.close()
    except Exception as e:
        result["error"] = str(e)
    return result


# ==============================================================================
# CODE COMPILER / RUNNER  (Windows + Linux/Mac compatible)
# ==============================================================================
import platform
IS_WINDOWS = platform.system() == "Windows"

def _find_python():
    """Find the correct Python executable on this OS."""
    import shutil
    for cmd in ["python", "python3"]:
        if shutil.which(cmd):
            return cmd
    return "python"   # last resort

def _find_node():
    import shutil
    for cmd in ["node", "node.exe"]:
        if shutil.which(cmd):
            return cmd
    return "node"

def _find_java():
    import shutil
    return "javac" if shutil.which("javac") else None

def _find_gcc():
    import shutil
    return "gcc" if shutil.which("gcc") else None

def _find_gpp():
    import shutil
    return "g++" if shutil.which("g++") else ("cl" if shutil.which("cl") else None)

# Detect once at startup
_PYTHON_CMD = _find_python()
_NODE_CMD   = _find_node()
print(f"[COMPILER] Python={_PYTHON_CMD}  Node={_NODE_CMD}  OS={platform.system()}")


def run_code(language, code, stdin_input=""):
    """Compile and run user code cross-platform (Windows + Linux/Mac)."""
    result = {"stdout": "", "stderr": "", "error": "", "timed_out": False}
    lang   = language.lower().strip()

    # ── Map aliases ───────────────────────────────────────────────────────────
    lang = {"py":"python","js":"javascript","c++":"cpp","node":"javascript"}.get(lang, lang)

    if lang not in ("python","java","javascript","c","cpp"):
        result["error"] = f"Unsupported language: {language}"
        return result

    try:
        with tempfile.TemporaryDirectory() as tmpdir:

            if lang == "python":
                fpath = os.path.join(tmpdir, "solution.py")
                with open(fpath, "w", encoding="utf-8") as f:
                    f.write(code)
                cmd  = [_PYTHON_CMD, fpath]
                proc = subprocess.run(cmd, input=stdin_input,
                                      capture_output=True, text=True,
                                      timeout=10)

            elif lang == "javascript":
                fpath = os.path.join(tmpdir, "solution.js")
                with open(fpath, "w", encoding="utf-8") as f:
                    f.write(code)
                cmd  = [_NODE_CMD, fpath]
                proc = subprocess.run(cmd, input=stdin_input,
                                      capture_output=True, text=True,
                                      timeout=10)

            elif lang == "java":
                fpath = os.path.join(tmpdir, "Main.java")
                with open(fpath, "w", encoding="utf-8") as f:
                    f.write(code)
                # Step 1: compile
                comp = subprocess.run(
                    ["javac", fpath],
                    capture_output=True, text=True, timeout=15
                )
                if comp.returncode != 0:
                    result["stderr"] = comp.stderr
                    result["error"]  = "Compilation error"
                    return result
                # Step 2: run (cwd=tmpdir so java finds Main.class)
                proc = subprocess.run(
                    ["java", "-cp", tmpdir, "Main"],
                    input=stdin_input, capture_output=True, text=True,
                    timeout=10
                )

            elif lang == "c":
                fpath  = os.path.join(tmpdir, "solution.c")
                outbin = os.path.join(tmpdir, "solution.exe" if IS_WINDOWS else "solution.out")
                with open(fpath, "w", encoding="utf-8") as f:
                    f.write(code)
                comp = subprocess.run(
                    ["gcc", fpath, "-o", outbin, "-lm"],
                    capture_output=True, text=True, timeout=15
                )
                if comp.returncode != 0:
                    result["stderr"] = comp.stderr
                    result["error"]  = "Compilation error"
                    return result
                proc = subprocess.run(
                    [outbin], input=stdin_input,
                    capture_output=True, text=True, timeout=10
                )

            elif lang == "cpp":
                fpath  = os.path.join(tmpdir, "solution.cpp")
                outbin = os.path.join(tmpdir, "solution.exe" if IS_WINDOWS else "solution.out")
                with open(fpath, "w", encoding="utf-8") as f:
                    f.write(code)
                comp = subprocess.run(
                    ["g++", fpath, "-o", outbin, "-std=c++17"],
                    capture_output=True, text=True, timeout=15
                )
                if comp.returncode != 0:
                    result["stderr"] = comp.stderr
                    result["error"]  = "Compilation error"
                    return result
                proc = subprocess.run(
                    [outbin], input=stdin_input,
                    capture_output=True, text=True, timeout=10
                )

            result["stdout"] = proc.stdout
            result["stderr"] = proc.stderr
            if proc.returncode != 0 and not proc.stdout and proc.stderr:
                result["error"] = proc.stderr[:300]

    except subprocess.TimeoutExpired:
        result["timed_out"] = True
        result["error"] = "Time limit exceeded (10s)"
    except FileNotFoundError as e:
        # Executable not found — give clear install instructions
        exe = str(e).split("'")[1] if "'" in str(e) else str(e)
        install_hints = {
            _PYTHON_CMD: "Install Python from https://python.org — tick 'Add Python to PATH'",
            "javac":     "Install JDK from https://adoptium.net — tick 'Set JAVA_HOME'",
            "java":      "Install JDK from https://adoptium.net",
            "gcc":       "Install MinGW-w64 from https://winlibs.com — add bin folder to PATH",
            "g++":       "Install MinGW-w64 from https://winlibs.com — add bin folder to PATH",
            _NODE_CMD:   "Install Node.js from https://nodejs.org",
        }
        hint = install_hints.get(exe, f"'{exe}' not found — install it and add to PATH")
        result["error"] = f"Compiler not found: {hint}"
    except Exception as e:
        result["error"] = str(e)
    return result


def judge_code(language, code, test_cases):
    """Run code against all test cases, return per-case pass/fail + partial score."""
    results = []
    passed  = 0
    for tc in test_cases:
        stdin    = tc["input"].replace("\\n", "\n")
        expected = tc["expected_output"].strip()
        run      = run_code(language, code, stdin)
        actual   = run["stdout"].strip()
        # Normalize line endings for comparison
        ok = (actual.replace("\r","") == expected.replace("\r",""))              and not run["timed_out"] and not run["error"]
        if ok: passed += 1
        results.append({
            "input":     tc["input"],
            "expected":  expected,
            "actual":    actual,
            "passed":    ok,
            "error":     (run.get("error","") or run.get("stderr",""))[:300],
            "timed_out": run.get("timed_out", False),
        })
    return {"test_results": results, "passed": passed, "total": len(test_cases)}


# ==============================================================================
# STUDENT ROUTES
# ==============================================================================
@app.route("/")
def index(): return redirect(url_for("student_login"))


def assign_questions_to_session(s):
    """Assign randomized question set to a session based on email."""
    seed = int(hashlib.md5(s["email"].encode()).hexdigest(),16) % (2**32)
    s["questions"] = get_random_questions_for_student(seed)
    s["mcq_total"]  = sum(q["marks"] for q in s["questions"]["mcq"])
    s["sql_total"]  = sum(q["marks"] for q in s["questions"]["sql"])
    s["code_total"] = sum(q["marks"] for q in s["questions"]["coding"])
    s["total_marks"]= s["mcq_total"] + s["sql_total"] + s["code_total"]


@app.route("/student/login", methods=["GET","POST"])
def student_login():
    token_error = request.args.get("error","")
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        u = str(data.get("username","")).strip()
        e = str(data.get("email","")).strip()
        if not u or not e:
            return jsonify({"error":"Username and email required"}), 400
        sid = new_session(u, e)
        s   = SESSIONS[sid]
        assign_questions_to_session(s)
        session["exam_sid"] = sid
        session["username"] = u
        return jsonify({"redirect": url_for("exam_page")})
    return render_template("student_login.html", error=token_error)


@app.route("/exam/start/<token>")
def exam_start_token(token):
    """Student arrives via email invite link."""
    invite = EXAM_INVITES.get(token)
    if not invite:
        return redirect(url_for("student_login", error="Invalid or expired exam link."))
    if invite.get("used"):
        return redirect(url_for("student_login", error="This exam link has already been used."))
    from datetime import datetime as DT
    if DT.fromisoformat(invite["expires"]) < DT.now():
        return redirect(url_for("student_login", error="This exam link has expired."))
    sid = new_session(invite["name"], invite["email"])
    s   = SESSIONS[sid]
    s["questions"]    = get_random_questions_for_student(invite["seed"])
    s["invite_token"] = token
    s["mcq_total"]  = sum(q["marks"] for q in s["questions"]["mcq"])
    s["sql_total"]  = sum(q["marks"] for q in s["questions"]["sql"])
    s["code_total"] = sum(q["marks"] for q in s["questions"]["coding"])
    s["total_marks"]= s["mcq_total"] + s["sql_total"] + s["code_total"]
    invite["used"]  = True
    session["exam_sid"] = sid
    session["username"] = invite["name"]
    return redirect(url_for("exam_page"))


@app.route("/exam")
@student_required
def exam_page():
    sid = session["exam_sid"]
    s   = SESSIONS.get(sid, {})
    if not s.get("questions"):
        assign_questions_to_session(s)
    yolo_info = {
        "loaded":    _yolo_model is not None,
        "finetuned": _yolo_fine,
        "phone_cls": YOLO_PHONE_ID,
        "person_cls":YOLO_PERSON_ID,
    }
    qs = s["questions"]
    return render_template("exam.html",
                           mcq_questions   = qs["mcq"],
                           sql_questions   = qs["sql"],
                           coding_questions= qs["coding"],
                           duration=EXAM_DURATION,
                           username=s.get("username",""),
                           session_id=sid,
                           yolo_info=yolo_info)


@app.route("/exam/submit", methods=["POST"])
@student_required
def submit_exam():
    sid  = session.get("exam_sid")
    data = request.get_json(silent=True) or {}
    s    = SESSIONS.get(sid)
    if not s: return jsonify({"error":"Session not found"}), 404
    if s["status"] == "submitted":
        return jsonify({"error":"Already submitted"}), 400

    qs = s.get("questions") or get_random_questions_for_student()

    # Score Part A — MCQ
    mcq_answers = data.get("answers", {})
    s["answers"]  = mcq_answers
    mcq_score = sum(q["marks"] for q in qs["mcq"]
                    if str(q["id"]) in mcq_answers
                    and int(mcq_answers[str(q["id"])]) == q["correct"])
    s["mcq_score"] = mcq_score

    # Score Part B — SQL (submitted as {qid: sql_string})
    sql_answers = data.get("sql_answers", {})
    s["sql_answers"] = sql_answers
    sql_score = 0
    sql_results = {}
    for q in qs["sql"]:
        user_sql = sql_answers.get(str(q["id"]),"").strip()
        if user_sql:
            res = run_sql_query(q["setup_sql"], user_sql, q["expected_output"], q["expected_cols"])
            sql_results[q["id"]] = res
            if res["passed"]:
                sql_score += q["marks"]
    s["sql_score"]   = sql_score
    s["sql_results"] = sql_results

    # Score Part C — Coding (submitted as {qid: {language, code}})
    code_answers = data.get("code_answers", {})
    s["code_answers"] = code_answers
    code_score = 0
    code_results = {}
    for q in qs["coding"]:
        sub = code_answers.get(str(q["id"]))
        if sub and sub.get("code","").strip():
            res = judge_code(sub.get("language","python"), sub["code"], q["test_cases"])
            code_results[q["id"]] = res
            # Partial marks: score proportional to test cases passed
            frac = res["passed"] / max(res["total"],1)
            code_score += int(q["marks"] * frac)
    s["code_score"]   = code_score
    s["code_results"] = code_results

    total_score = mcq_score + sql_score + code_score
    s["score"]    = total_score
    s["end_time"] = datetime.now().isoformat()
    s["status"]   = "submitted"
    s["cheating_score"] = cheating_score(s)
    session.pop("exam_sid", None)

    # Send result email in background thread
    def send_bg():
        try:
            send_result_email(s["username"], s["email"], s, qs)
        except Exception as ex:
            print(f"[EMAIL] Result send failed: {ex}")
    threading.Thread(target=send_bg, daemon=True).start()

    return jsonify({"score":total_score,"total":s["total_marks"],
                    "cheating_score":s["cheating_score"],
                    "redirect":url_for("exam_result",sid=sid)})


@app.route("/exam/result/<sid>")
def exam_result(sid):
    s = SESSIONS.get(sid)
    if not s: return redirect(url_for("student_login"))
    qs = s.get("questions") or {"mcq": QUESTIONS[:5], "sql": [], "coding": []}
    return render_template("result.html", session=s,
                           mcq_questions    = qs["mcq"],
                           sql_questions    = qs["sql"],
                           coding_questions = qs["coding"])


# ==============================================================================
# DETECTION APIs
# ==============================================================================
@app.route("/api/enrol", methods=["POST"])
def enrol():
    sid = session.get("exam_sid"); s = SESSIONS.get(sid) if sid else None
    if not s: return jsonify({"ok":False,"error":"No session"}), 400
    data = request.get_json(silent=True) or {}
    if not s.get("ref_embed"): s["ref_embed"] = {}

    if data.get("voice_sample"):
        s["ref_embed"]["audio_rms"] = float(data.get("audio_rms",0.05))
        s["ref_embed"]["audio_zcr"] = float(data.get("audio_zcr",0.30))
        s["enrolled"] = True
        print(f"[ENROL] Voice: {s['username']}  rms={s['ref_embed']['audio_rms']:.4f}")
        return jsonify({"ok":True,"step":"voice"})

    b64 = data.get("frame","")
    if not b64: return jsonify({"ok":False,"error":"No frame"}), 400
    try:
        if "," in b64: b64 = b64.split(",",1)[1]
        img_bytes = base64.b64decode(b64+"==")
        if HAS_CV:
            nparr = np.frombuffer(img_bytes,np.uint8)
            img   = cv2.imdecode(nparr,cv2.IMREAD_COLOR)
            if img is not None:
                gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
                casc = cv2.CascadeClassifier(
                    cv2.data.haarcascades+"haarcascade_frontalface_default.xml")
                fcs  = casc.detectMultiScale(gray,1.1,4,minSize=(60,60))
                if len(fcs) > 0:
                    x,y,fw,fh = fcs[0]
                    roi = gray[y:y+fh,x:x+fw].astype(np.float32)/255.0
                    s["ref_embed"]["face_mean"] = float(roi.mean())
                    s["ref_embed"]["face_std"]  = float(roi.std())
                    print(f"[ENROL] Face detected: {s['username']}")
                else:
                    gf = gray.astype(np.float32)/255.0
                    s["ref_embed"]["face_mean"] = float(gf.mean())
                    s["ref_embed"]["face_std"]  = float(gf.std())
        else:
            raw = np.frombuffer(img_bytes,np.uint8)
            s["ref_embed"]["face_mean"] = float(raw.mean()/255.0)
            s["ref_embed"]["face_std"]  = 0.15
        return jsonify({"ok":True,"step":"face"})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)}), 500


@app.route("/api/detect_frame", methods=["POST"])
def detect_frame():
    sid  = session.get("exam_sid"); s = SESSIONS.get(sid) if sid else None
    data = request.get_json(silent=True) or {}
    b64  = data.get("frame","")

    try:
        if "," in b64: b64 = b64.split(",",1)[1]
        img_bytes  = base64.b64decode(b64+"==")
        audio_rms  = float(data.get("audio_rms",  0.0))
        audio_zcr  = float(data.get("audio_zcr",  0.0))
        audio_peak = float(data.get("audio_peak", 0.0))
        ref_embed  = s.get("ref_embed") if s else None

        feat, phone_boxes, person_boxes, person_count = \
            extract_features(img_bytes, audio_rms, audio_zcr, audio_peak, ref_embed)

        result = predict_ensemble(feat)
        cls    = result["pred_class"]

        # ── Alert flags ────────────────────────────────────────────────────────
        # Phone: YOLO conf >= 0.15 (COCO class 67 gives low scores) OR CNN says phone
        yolo_phone = float(feat[0]) >= 0.15
        # Multi-person: YOLO must find 2+ real person boxes with low overlap between them
        # Also require 2nd box confidence >= 0.65 (same as our person filter threshold)
        yolo_multi = (float(feat[16]) >= 2.0 and float(feat[15]) >= 0.65
                      and _boxes_non_overlapping(person_boxes))
        # Face mismatch — voice identity REMOVED (unreliable), face only
        face_mismatch_raw = (float(feat[37]) < 0.40 and
                             (s.get("ref_embed") if s else None) is not None)

        phone_alert    = (cls == 1) or yolo_phone
        multi_alert    = yolo_multi                      # YOLO only — CNN alone causes false positives
        # voice_alert: CNN class 3 OR audio RMS above 0.08 (lowered from 0.18)
        # feat[27] threshold lowered to 0.25 (was 0.55) — normal speech ~0.10-0.20 RMS
        speech_flag    = bool(data.get("speech_detected", False))
        voice_alert    = (cls == 3) or (audio_rms > 0.06 and feat[27] > 0.20) or speech_flag
        identity_alert = face_mismatch_raw               # face-only identity check
        side_alert     = bool(feat[55] > 0.45)

        # ── Temporal smoothing ─────────────────────────────────────────────────
        sk = sid or "anon"
        if sk not in _smooth_buf:
            _smooth_buf[sk] = {"phone":0,"multi":0,"voice":0,"identity":0}
        buf = _smooth_buf[sk]
        # Phone is sticky — never decrements (student may briefly hide phone)
        buf["phone"]    = (buf["phone"] + 1)    if phone_alert    else buf["phone"]
        buf["multi"]    = (buf["multi"]    + 1) if multi_alert    else max(0, buf["multi"]    - 1)
        buf["voice"]    = (buf["voice"]    + 1) if voice_alert    else max(0, buf["voice"]    - 1)
        buf["identity"] = (buf["identity"] + 1) if identity_alert else max(0, buf["identity"] - 1)

        pc = buf["phone"]    >= _SMOOTH["phone"]
        mc = buf["multi"]    >= _SMOOTH["multi"]
        vc = buf["voice"]    >= _SMOOTH["voice"]
        ic = buf["identity"] >= _SMOOTH["identity"]

        # ── Count violations ───────────────────────────────────────────────────
        violations = []; voice_msg = ""
        if pc: violations.append("phone");         voice_msg = "Warning! Do not use your phone."
        if mc: violations.append("multi");         voice_msg = "Warning! Only one person is allowed."
        if vc: violations.append("voice");         voice_msg = "Warning! Please remain silent."
        if ic: violations.append("face_mismatch"); voice_msg = "Warning! Face does not match enrolled student."

        if s and violations:
            s["warnings"].append({"time":datetime.now().isoformat(),"types":violations})
            if pc: s["phone_count"]          += 1
            if mc: s["multi_person_count"]   += 1
            if vc: s["voice_count"]          += 1
            if ic: s["face_mismatch_count"]   = s.get("face_mismatch_count",0) + 1
            s["cheating_score"] = cheating_score(s)
        if s and side_alert:
            s["side_view_count"] = s.get("side_view_count",0) + 1

        # ── Auto-termination ───────────────────────────────────────────────────
        terminate        = False
        terminate_reason = ""
        terminate_cause  = ""
        if s and not s.get("terminated"):
            ph_n = s.get("phone_count",         0)
            mu_n = s.get("multi_person_count",   0)
            fm_n = s.get("face_mismatch_count",  0)
            if ph_n >= PHONE_TERMINATE:
                terminate       = True
                terminate_cause = "phone"
                terminate_reason = (
                    f"📵 Phone detected {ph_n} times — "
                    f"Mobile devices are strictly prohibited during the exam."
                )
            elif mu_n >= MULTI_TERMINATE:
                terminate       = True
                terminate_cause = "multi_person"
                terminate_reason = (
                    f"👥 Multiple persons detected {mu_n} times — "
                    f"Only the registered student is allowed in the exam room."
                )
            elif fm_n >= FACE_MISMATCH_TERMINATE:
                terminate       = True
                terminate_cause = "face_mismatch"
                terminate_reason = (
                    f"🪪 Face mismatch detected {fm_n} times — "
                    f"The person in frame does not match the enrolled student."
                )
            if terminate:
                s["terminated"]       = True
                s["terminate_reason"] = terminate_reason
                s["terminate_cause"]  = terminate_cause
                s["terminate_time"]   = datetime.now().isoformat()
                s["end_time"]         = datetime.now().isoformat()
                s["status"]           = "terminated"

        return jsonify({
            "verdict":        result["class_name"],
            "pred_class":     cls,
            "confidence":     result["confidence"],
            "probabilities":  result["probabilities"],
            # Alert flags (smoothed)
            "phone_alert":    pc,
            "multi_alert":    mc,
            "voice_alert":    vc,
            "face_mismatch":  ic,
            "identity_alert": ic,
            "side_view":      side_alert,
            # Raw YOLO results for canvas overlay
            "phone_boxes":    [[round(v,4) for v in b] for b in phone_boxes],
            "person_boxes":   [[round(v,4) for v in b] for b in person_boxes],
            "person_count":   person_count,
            "yolo_phone_conf":round(float(feat[0]),3),
            "yolo_finetuned": _yolo_fine,
            # Identity scores
            "face_match":    round(float(feat[37]),3),
            "voice_match":   round(float(feat[41]),3),
            "identity_conf": round(float(feat[43]),3),
            # Head pose
            "head_yaw":   round(float(feat[48]),3),
            "side_score": round(float(feat[55]),3),
            # Live counts for frontend violation counters
            "phone_count": s.get("phone_count",0)         if s else 0,
            "multi_count": s.get("multi_person_count",0)  if s else 0,
            "face_count":  s.get("face_mismatch_count",0) if s else 0,
            "tab_count":   s.get("tab_switch_count",0)    if s else 0,
            # Termination info
            "terminate":        terminate,
            "redirect_url":     f"/exam/result/{sid}" if terminate else "",
            "terminate_reason": terminate_reason,
            "terminate_cause":  terminate_cause,
            # Session
            "violations":     violations,
            "voice_message":  voice_msg,
            "cheating_score": s["cheating_score"] if s else 0,
            "warning_count":  len(s["warnings"])  if s else 0,
        })

    except Exception as e:
        print(f"[detect_frame] {traceback.format_exc()}")
        return jsonify({"verdict":"Normal","pred_class":0,"confidence":0.9,
                        "phone_alert":False,"multi_alert":False,"voice_alert":False,
                        "identity_alert":False,"side_view":False,
                        "phone_boxes":[],"person_boxes":[],"person_count":1,
                        "yolo_phone_conf":0.0,"face_match":0.9,
                        "identity_conf":0.9,
                        "phone_count":0,"multi_count":0,"face_count":0,"tab_count":0,
                        "terminate":False,"terminate_reason":"","terminate_cause":"",
                        "violations":[],"voice_message":"",
                        "cheating_score":0,"warning_count":0,"error":str(e)})


@app.route("/api/transcript", methods=["POST"])
def save_transcript():
    sid = session.get("exam_sid"); s = SESSIONS.get(sid) if sid else None
    if not s: return jsonify({"ok":False}), 400
    data = request.get_json(silent=True) or {}
    # Accept both "text" (sent by exam.html) and legacy "txt" key
    txt = str(data.get("text", data.get("txt", ""))).strip()
    conf = float(data.get("confidence", 0.0))
    if txt:
        s["transcript"].append({
            "text": txt, "txt": txt,
            "time": datetime.now().isoformat(),
            "confidence": conf,
        })
        print(f"[TRANSCRIPT] {s['username']} conf={conf:.2f}: {txt[:60]}")
        # Directly flag voice violation when speech confirmed
        if conf > 0.30 or len(txt.split()) >= 2:
            sk = session.get("exam_sid","") or ""
            if sk not in _smooth_buf:
                _smooth_buf[sk] = {"phone":0,"multi":0,"voice":0,"identity":0}
            _smooth_buf[sk]["voice"] = max(_smooth_buf[sk]["voice"], _SMOOTH["voice"])
            s["voice_count"] = s.get("voice_count", 0) + 1
            s["cheating_score"] = cheating_score(s)
            print(f"[VOICE] {s['username']} confirmed — {txt[:40]}")
    return jsonify({"ok":True,"segments":len(s["transcript"])})


@app.route("/api/tab_switch", methods=["POST"])
def tab_switch():
    sid = session.get("exam_sid"); s = SESSIONS.get(sid) if sid else None
    terminate = False; terminate_reason = ""; terminate_cause = ""
    if s and not s.get("terminated"):
        s["tab_switch_count"] += 1
        s["warnings"].append({"time":datetime.now().isoformat(),"types":["tab_switch"]})
        s["cheating_score"] = cheating_score(s)
        tab_n = s["tab_switch_count"]
        if tab_n >= TAB_TERMINATE:
            terminate        = True
            terminate_cause  = "tab_switch"
            terminate_reason = (
                f"🔀 Tab switched {tab_n} times — "
                f"Navigating away from the exam is strictly prohibited."
            )
            s["terminated"]       = True
            s["terminate_reason"] = terminate_reason
            s["terminate_cause"]  = terminate_cause
            s["terminate_time"]   = datetime.now().isoformat()
            s["end_time"]         = datetime.now().isoformat()
            s["status"]           = "terminated"
    return jsonify({
        "tab_switches":     s["tab_switch_count"] if s else 0,
        "cheating_score":   s["cheating_score"]   if s else 0,
        "terminate":        terminate,
        "terminate_reason": terminate_reason,
        "terminate_cause":  terminate_cause,
    })


@app.route("/api/screenshot", methods=["POST"])
def save_screenshot():
    sid  = session.get("exam_sid"); s = SESSIONS.get(sid) if sid else None
    data = request.get_json(silent=True) or {}
    # Accept both "frame" (sent by exam.html) and "image" (legacy)
    b64  = data.get("frame", data.get("image", ""))
    reason = str(data.get("reason","violation"))[:40]
    if not b64 or not s: return jsonify({"saved":False,"error":"no data or session"})
    try:
        if "," in b64: b64 = b64.split(",",1)[1]
        # Pad base64 string to valid length
        b64 = b64 + "=" * (-len(b64) % 4)
        img = base64.b64decode(b64)
        fn  = f"{sid[:8]}_{int(time.time())}_{reason}.jpg"
        fpath = os.path.join(SS_DIR, fn)
        with open(fpath, "wb") as f: f.write(img)
        s["screenshots"].append({
            "file":   fn,
            "reason": reason,
            "time":   datetime.now().isoformat(),
            "size":   len(img),
        })
        print(f"[SS] Saved {fn} ({len(img)} bytes) reason={reason}")
        return jsonify({"saved":True,"file":fn})
    except Exception as e:
        print(f"[SS] Error: {e}")
        return jsonify({"saved":False,"error":str(e)})


@app.route("/api/model_status")
def model_status():
    return jsonify({
        "model_type":    _model_type,
        "cnn_loaded":    _cnn_model  is not None,
        "rf_loaded":     _rf_model   is not None,
        "scaler_loaded": _scaler     is not None,
        "yolo_loaded":   _yolo_model is not None,
        "yolo_finetuned":_yolo_fine,
        "yolo_phone_class": YOLO_PHONE_ID,
        "yolo_person_class":YOLO_PERSON_ID,
        "has_cv2":  HAS_CV, "has_tf":HAS_TF, "has_yolo":HAS_YOLO,
        "num_features":  NUM_FEATURES,
        "class_names":   CLASS_NAMES,
        "version": "v6",
    })



# ==============================================================================
# COMPILER + SQL + INVITE APIs
# ==============================================================================
@app.route("/api/run_code", methods=["POST"])
@student_required
def api_run_code():
    """Run code for a given question and input. Returns stdout/stderr."""
    data     = request.get_json(silent=True) or {}
    language = data.get("language", "python")
    code     = data.get("code", "")
    stdin    = data.get("stdin", "")
    if not code.strip():
        return jsonify({"error": "No code provided"}), 400
    result = run_code(language, code, stdin)
    return jsonify(result)


@app.route("/api/run_tests", methods=["POST"])
@student_required
def api_run_tests():
    """Judge code against all test cases for a question."""
    data     = request.get_json(silent=True) or {}
    sid      = session.get("exam_sid")
    s        = SESSIONS.get(sid)
    language = data.get("language", "python")
    code     = data.get("code", "")
    qid      = str(data.get("question_id", ""))
    if not s: return jsonify({"error": "No session"}), 400
    # Find the question
    qs = s.get("questions") or {}
    q  = next((q for q in qs.get("coding", []) if str(q["id"]) == qid), None)
    if not q: return jsonify({"error": "Question not found"}), 404
    result = judge_code(language, code, q["test_cases"])
    return jsonify(result)


@app.route("/api/test_sql", methods=["POST"])
@student_required
def api_test_sql():
    """Test a student's SQL query (non-scored trial run)."""
    data  = request.get_json(silent=True) or {}
    sid   = session.get("exam_sid")
    s     = SESSIONS.get(sid)
    qid   = str(data.get("question_id", ""))
    query = data.get("query", "").strip()
    if not s: return jsonify({"error": "No session"}), 400
    qs = s.get("questions") or {}
    q  = next((q for q in qs.get("sql", []) if str(q["id"]) == qid), None)
    if not q: return jsonify({"error": "Question not found"}), 404
    result = run_sql_query(q["setup_sql"], query, q["expected_output"], q["expected_cols"])
    # Don't reveal whether it passed — only show output
    result.pop("expected", None)
    return jsonify(result)


# ── Admin: Send exam invite ────────────────────────────────────────────────────
@app.route("/admin/api/send_invite", methods=["POST"])
@admin_required
def admin_send_invite():
    """Send exam invite email to one or many students."""
    data      = request.get_json(silent=True) or {}
    students  = data.get("students", [])  # [{name, email}, ...]
    if not students:
        return jsonify({"error": "No students provided"}), 400
    results = []
    from datetime import datetime as DT, timedelta
    for st in students:
        name  = str(st.get("name","")).strip()
        email = str(st.get("email","")).strip()
        if not name or not email:
            results.append({"email": email, "ok": False, "error": "Missing name/email"})
            continue
        token   = secrets.token_urlsafe(32)
        seed    = int(hashlib.md5(email.encode()).hexdigest(), 16) % (2**32)
        expires = (DT.now() + timedelta(hours=EXAM_LINK_EXPIRY)).isoformat()
        EXAM_INVITES[token] = {
            "name": name, "email": email, "seed": seed,
            "expires": expires, "used": False,
            "created": DT.now().isoformat(),
            "token": token,
        }
        ok, err = send_exam_invite_email(name, email, token)
        results.append({"name": name, "email": email, "ok": ok,
                        "error": err, "token": token,
                        "link": f"{EXAM_BASE_URL}/exam/start/{token}"})
    sent = sum(1 for r in results if r["ok"])
    return jsonify({"sent": sent, "total": len(results), "results": results})


@app.route("/admin/api/invites")
@admin_required
def admin_list_invites():
    """List all sent exam invites."""
    rows = sorted(EXAM_INVITES.values(),
                  key=lambda x: x.get("created",""), reverse=True)
    return jsonify(rows)


@app.route("/admin/api/question_bank")
@admin_required
def admin_question_bank():
    """Return question bank stats for admin."""
    return jsonify({
        "mcq_total":    len(MCQ_BANK),
        "sql_total":    len(SQL_BANK),
        "coding_total": len(CODING_BANK),
        "mcq_per_exam": MCQ_PER_EXAM,
        "sql_per_exam": SQL_PER_EXAM,
        "code_per_exam":CODE_PER_EXAM,
        "mcq":    MCQ_BANK,   # includes section and topic fields
        "sql":    [{k:v for k,v in q.items() if k!="setup_sql"} for q in SQL_BANK],
        "coding": [{k:v for k,v in q.items() if k not in ("template_python","template_java")} for q in CODING_BANK],
    })



# ── Admin: Email settings ──────────────────────────────────────────────────────
@app.route("/admin/api/email_config", methods=["GET","POST"])
@admin_required
def admin_email_config():
    if request.method == "GET":
        safe = {k:v for k,v in EMAIL_CONFIG.items() if k not in ("smtp_pass","api_key")}
        safe["smtp_pass"] = "***" if EMAIL_CONFIG.get("smtp_pass") else ""
        safe["api_key"]   = "***" if EMAIL_CONFIG.get("api_key")   else ""
        return jsonify({"config": safe, "presets": PROVIDER_PRESETS})
    data = request.get_json(silent=True) or {}
    EMAIL_CONFIG["provider"]   = data.get("provider",   EMAIL_CONFIG["provider"])
    EMAIL_CONFIG["smtp_host"]  = data.get("smtp_host",  EMAIL_CONFIG["smtp_host"])
    EMAIL_CONFIG["smtp_port"]  = int(data.get("smtp_port", EMAIL_CONFIG["smtp_port"]))
    EMAIL_CONFIG["smtp_user"]  = data.get("smtp_user",  EMAIL_CONFIG["smtp_user"])
    EMAIL_CONFIG["from_name"]  = data.get("from_name",  EMAIL_CONFIG.get("from_name","ProctorAI Exams"))
    EMAIL_CONFIG["from_email"] = data.get("from_email", EMAIL_CONFIG.get("from_email",""))
    EMAIL_CONFIG["use_ssl"]    = bool(data.get("use_ssl", EMAIL_CONFIG["use_ssl"]))
    EMAIL_CONFIG["enabled"]    = bool(data.get("enabled", EMAIL_CONFIG["enabled"]))
    if data.get("smtp_pass","") not in ("", "***"):
        EMAIL_CONFIG["smtp_pass"] = data["smtp_pass"]
    if data.get("api_key","") not in ("", "***"):
        EMAIL_CONFIG["api_key"] = data["api_key"]
    safe = {k:v for k,v in EMAIL_CONFIG.items() if k not in ("smtp_pass","api_key")}
    return jsonify({"ok": True, "config": safe})


@app.route("/admin/api/test_email", methods=["POST"])
@admin_required
def admin_test_email():
    data    = request.get_json(silent=True) or {}
    test_to = data.get("to", EMAIL_CONFIG.get("smtp_user",""))
    ok, msg = test_email_connection()
    if ok:
        ok2, msg2 = send_email(test_to,
            "ProctorAI — SMTP Test",
            "<h2>✅ SMTP is working!</h2><p>Your ProctorAI email setup is correct.</p>")
        return jsonify({"ok": ok2, "message": msg2 or msg})
    return jsonify({"ok": False, "message": msg})


@app.route("/admin/api/exam_settings", methods=["GET","POST"])
@admin_required
def admin_exam_settings():
    global MCQ_PER_EXAM, SQL_PER_EXAM, CODE_PER_EXAM, EXAM_DURATION, EXAM_BASE_URL
    if request.method == "GET":
        return jsonify({"mcq_per_exam":MCQ_PER_EXAM,"sql_per_exam":SQL_PER_EXAM,
                        "code_per_exam":CODE_PER_EXAM,"exam_duration":EXAM_DURATION,
                        "exam_base_url":EXAM_BASE_URL})
    data = request.get_json(silent=True) or {}
    if "mcq_per_exam"  in data: MCQ_PER_EXAM  = int(data["mcq_per_exam"])
    if "sql_per_exam"  in data: SQL_PER_EXAM  = int(data["sql_per_exam"])
    if "code_per_exam" in data: CODE_PER_EXAM = int(data["code_per_exam"])
    if "exam_duration" in data: EXAM_DURATION = int(data["exam_duration"])
    if "exam_base_url" in data: EXAM_BASE_URL = str(data["exam_base_url"]).strip()
    return jsonify({"ok":True})


# ── Admin: Question CRUD ───────────────────────────────────────────────────────
@app.route("/admin/api/question", methods=["POST"])
@admin_required
def admin_add_question():
    """Add a new question to a bank. body: {bank, question_data}"""
    data  = request.get_json(silent=True) or {}
    bank  = data.get("bank")   # mcq | sql | coding
    qdata = data.get("question", {})
    if bank == "mcq":
        qdata["id"] = f"m{len(MCQ_BANK)+1}_{int(time.time())}"
        qdata.setdefault("marks", 2)
        MCQ_BANK.append(qdata)
    elif bank == "sql":
        qdata["id"] = f"s{len(SQL_BANK)+1}_{int(time.time())}"
        qdata.setdefault("marks", 5)
        SQL_BANK.append(qdata)
    elif bank == "coding":
        qdata["id"] = f"c{len(CODING_BANK)+1}_{int(time.time())}"
        qdata.setdefault("marks", 10)
        qdata.setdefault("test_cases", [])
        CODING_BANK.append(qdata)
    else:
        return jsonify({"error": "Invalid bank"}), 400
    return jsonify({"ok": True, "id": qdata["id"]})


@app.route("/admin/api/question/<qid>", methods=["PUT","DELETE"])
@admin_required
def admin_edit_question(qid):
    """Edit or delete a question by id."""
    banks = {"m": MCQ_BANK, "s": SQL_BANK, "c": CODING_BANK}
    prefix = qid[0] if qid else ""
    bank = banks.get(prefix)
    if bank is None:
        return jsonify({"error":"Unknown question prefix"}), 400
    idx = next((i for i,q in enumerate(bank) if str(q["id"]) == qid), None)
    if idx is None:
        return jsonify({"error":"Question not found"}), 404
    if request.method == "DELETE":
        bank.pop(idx)
        return jsonify({"ok": True})
    data = request.get_json(silent=True) or {}
    bank[idx].update(data)
    bank[idx]["id"] = qid  # ensure id unchanged
    return jsonify({"ok": True})


@app.route("/admin/api/question/<qid>/test_case", methods=["POST"])
@admin_required
def admin_add_test_case(qid):
    """Add a test case to a coding question."""
    q = next((q for q in CODING_BANK if str(q["id"]) == qid), None)
    if not q: return jsonify({"error":"Not found"}), 404
    tc = request.get_json(silent=True) or {}
    q.setdefault("test_cases",[]).append(tc)
    return jsonify({"ok":True,"count":len(q["test_cases"])})

# ==============================================================================
# ADMIN ROUTES
# ==============================================================================
@app.route("/admin/login", methods=["GET","POST"])
def admin_login():
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        u,p  = data.get("username",""), data.get("password","")
        if u in ADMINS and check_password_hash(ADMINS[u],p):
            session["admin_logged_in"] = True
            session["admin_user"] = u
            return jsonify({"redirect":url_for("admin_dashboard")})
        return jsonify({"error":"Invalid credentials"}), 401
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.clear(); return redirect(url_for("admin_login"))


@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    return render_template("admin_dashboard.html",
                           admin_user=session.get("admin_user","admin"))


@app.route("/admin/api/sessions")
@admin_required
def api_sessions():
    q    = request.args.get("search","").lower().strip()
    # Auto-expire sessions that have passed their time limit
    for s in SESSIONS.values():
        if session_expired(s):
            s["status"] = "expired"
            s["end_time"] = s.get("end_time") or datetime.now().isoformat()
    rows = sorted(SESSIONS.values(),key=lambda x:x.get("start_time",""),reverse=True)
    if q: rows = [s for s in rows if q in s["username"].lower() or q in s["email"].lower()]
    return jsonify([{
        "session_id":          s["session_id"],
        "username":            s["username"],
        "email":               s["email"],
        "start_time":          s.get("start_time",""),
        "end_time":            s.get("end_time"),
        "status":              s.get("status","active"),
        "score":               s.get("score"),
        "mcq_score":           s.get("mcq_score",0),
        "sql_score":           s.get("sql_score",0),
        "code_score":          s.get("code_score",0),
        "mcq_total":           s.get("mcq_total",0),
        "sql_total":           s.get("sql_total",0),
        "code_total":          s.get("code_total",0),
        "total_marks":         s.get("total_marks",0),
        "cheating_score":      s.get("cheating_score",0),
        "warning_count":       len(s.get("warnings",[])),
        "phone_count":         s.get("phone_count",0),
        "multi_person_count":  s.get("multi_person_count",0),
        "face_mismatch_count": s.get("face_mismatch_count",0),
        "voice_count":         s.get("voice_count",0),
        "side_view_count":     s.get("side_view_count",0),
        "tab_switch_count":    s.get("tab_switch_count",0),
        "screenshot_count":    len(s.get("screenshots",[])),
        "enrolled":            s.get("enrolled",False),
        "terminated":          s.get("terminated",False),
        "terminate_cause":     s.get("terminate_cause",""),
        "terminate_reason":    s.get("terminate_reason",""),
    } for s in rows])


@app.route("/admin/api/session/<sid>")
@admin_required
def api_session_detail(sid):
    s = SESSIONS.get(sid)
    if not s: return jsonify({"error":"Not found"}), 404
    return jsonify({k:v for k,v in s.items() if k != "ref_embed"})


@app.route("/admin/api/stats")
@admin_required
def api_stats():
    total   = len(SESSIONS)
    active  = sum(1 for s in SESSIONS.values() if s.get("status")=="active")
    flagged = sum(1 for s in SESSIONS.values() if s.get("cheating_score",0)>40)
    id_flag = sum(1 for s in SESSIONS.values() if s.get("face_mismatch_count",0)>0)
    term_count = sum(1 for s in SESSIONS.values() if s.get("terminated",False))
    scored  = [s["score"] for s in SESSIONS.values() if s.get("score") is not None]
    avg     = round(sum(scored)/len(scored),1) if scored else 0
    return jsonify({"total":total,"active":active,"submitted":total-active,
                    "flagged":flagged,"terminated":term_count,
                    "face_mismatch_flagged":id_flag,"avg_score":avg,
                    "model_type":_model_type,"yolo_active":_yolo_model is not None,
                    "yolo_finetuned":_yolo_fine})


@app.route("/admin/api/export")
@admin_required
def export_excel():
    rows = sorted(SESSIONS.values(),key=lambda x:x.get("start_time",""),reverse=True)
    base = f"exam_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if HAS_XL:
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Results"
        hf=PatternFill("solid",fgColor="003366"); hfn=Font(color="FFFFFF",bold=True)
        ts=Side(style="thin"); thin=Border(left=ts,right=ts,top=ts,bottom=ts)
        ff=PatternFill("solid",fgColor="FFE0E0")
        hdrs=["#","Username","Email","Started","Ended","Status","Score","Total",
              "Cheat%","Warnings","Phone","Multi","Face Mismatch","Voice","Side","Tab","Screenshots","Enrolled","Terminated"]
        wds=[4,16,26,20,20,11,8,7,8,9,8,8,13,8,8,8,12,9,11]
        for ci,(h,w) in enumerate(zip(hdrs,wds),1):
            c=ws.cell(1,ci,h); c.fill=hf; c.font=hfn; c.border=thin
            c.alignment=Alignment(horizontal="center")
            ws.column_dimensions[c.column_letter].width=w
        for ri,s in enumerate(rows,2):
            fl=s.get("cheating_score",0)>40
            vals=[ri-1,s["username"],s["email"],
                  (s.get("start_time") or "")[:19].replace("T"," "),
                  (s.get("end_time")   or "")[:19].replace("T"," "),
                  s.get("status",""),s.get("score","N/A"),s.get("total_marks",20),
                  s.get("cheating_score",0),len(s.get("warnings",[])),
                  s.get("phone_count",0),s.get("multi_person_count",0),
                  s.get("face_mismatch_count",0),
                  s.get("voice_count",0),
                  s.get("side_view_count",0),s.get("tab_switch_count",0),
                  len(s.get("screenshots",[])),
                  "Yes" if s.get("enrolled") else "No",
                  "Yes" if s.get("terminated") else "No"]
            for ci,v in enumerate(vals,1):
                c=ws.cell(ri,ci,v); c.border=thin
                c.alignment=Alignment(horizontal="center")
                if fl: c.fill=ff
        fn=base+".xlsx"; fp=os.path.join(EXPORT_DIR,fn)
        wb.save(fp); return send_file(fp,as_attachment=True,download_name=fn)

    fn=base+".csv"; fp=os.path.join(EXPORT_DIR,fn)
    with open(fp,"w",newline="",encoding="utf-8") as f:
        w=csv.writer(f)
        w.writerow(["Username","Email","Started","Ended","Status","Score","Total",
                    "Cheat%","Phone","Multi","FaceMismatch","Voice","SideView","Tab","Terminated"])
        for s in rows:
            w.writerow([s["username"],s["email"],
                        (s.get("start_time") or "")[:19],
                        (s.get("end_time")   or "")[:19],
                        s.get("status",""),s.get("score",""),s.get("total_marks",20),
                        s.get("cheating_score",0),s.get("phone_count",0),
                        s.get("multi_person_count",0),s.get("face_mismatch_count",0),
                        s.get("voice_count",0),s.get("side_view_count",0),
                        s.get("tab_switch_count",0),
                        "Yes" if s.get("terminated") else "No"])
    return send_file(fp,as_attachment=True,download_name=fn)


@app.route("/admin/screenshot/<fname>")
@admin_required
def get_screenshot(fname):
    return send_from_directory(SS_DIR,fname)


@app.route("/static/<path:filename>")
def static_files(filename):
    return send_from_directory("static",filename)


# ==============================================================================
# MAIN
# ==============================================================================
if __name__ == "__main__":
    print("\n"+"="*64)
    print("  Cognizant ProctorAI v6  —  COCO YOLOv8 Edition")
    print("  YOLOv8: Phone (class 67/1) + Person (class 0)")
    print("  CNN: 5-class (Normal|Phone|Multi|Voice|Identity)")
    print("="*64)
    load_models()
    print(f"\n  Model     : {_model_type.upper()}")
    print(f"  CNN (h5)  : {'✅ cnn_phone.h5' if _cnn_model else '⚠  not found — run training.py'}")
    print(f"  YOLOv8    : {'✅ Fine-tuned yolo_proctor.pt' if _yolo_fine else ('✅ COCO yolov8n.pt' if _yolo_model else '⚠  not loaded')}")
    print(f"  OpenCV    : {'✅' if HAS_CV else '⚠  pip install opencv-python'}")
    print(f"\n  Student → http://127.0.0.1:{PORT}/student/login")
    print(f"  Admin   → http://127.0.0.1:{PORT}/admin/login   (admin/admin123)")
    print("="*64+"\n")
    app.run(host="127.0.0.1", port=PORT, debug=True, use_reloader=False)